import openpyxl
import os

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class ScoreLineSkill:
    @staticmethod
    def get_batch_lines(year=None, category=None):
        """获取批次线数据"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/批次线-2021-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if year and len(row) > 1 and str(year) != str(row[1]):
                    continue
                if category and len(row) > 3 and category not in str(row[3]):
                    continue
                data.append(row)
        return data[:30]
    
    @staticmethod
    def get_rank_by_score(score, year=2024):
        """根据分数获取全省排名"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/一分一段-2017-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            if row[0] == score or (isinstance(row[0], (int, float)) and int(row[0]) == score):
                return row
        return None
    
    @staticmethod
    def get_school_scores(school_name=None, year=None, min_score=None):
        """获取院校分数线"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/院校分数-2020-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                if year and str(year) not in str(row):
                    continue
                if min_score and isinstance(row[-1], (int, float)) and row[-1] < min_score:
                    continue
                data.append(row)
                if len(data) >= 20:
                    break
        return data
    
    @staticmethod
    def get_major_scores(major_name=None, year=None, school_name=None):
        """获取专业分数线"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/专业分数-2024-考试院.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                if major_name and major_name not in str(row):
                    continue
                data.append(row)
                if len(data) >= 20:
                    break
        return data


class EnrollmentPlanSkill:
    @staticmethod
    def get_plan(year=None, school_name=None, major_name=None):
        """获取招生计划"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/招生计划/招生计划-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                if major_name and major_name not in str(row):
                    continue
                data.append(row)
                if len(data) >= 20:
                    break
        return data
    
    @staticmethod
    def get_plan_by_year(year):
        """按年份获取招生计划"""
        year_map = {
            2021: '招生计划-2021.xlsx',
            2022: '招生计划-2022.xlsx',
            2023: '招生计划-2023-本科.xlsx',
            2024: '招生计划-2024.xlsx',
            2025: '招生计划-2025.xlsx'
        }
        if year not in year_map:
            return []
        file_path = os.path.join(DATA_DIR, f'1_最近年份数据/招生计划/{year_map[year]}')
        if not os.path.exists(file_path):
            return []
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))[:30]


class SchoolInfoSkill:
    @staticmethod
    def get_school_info(school_name=None):
        """获取院校基础信息"""
        file_path = os.path.join(DATA_DIR, '3_院校信息/院校基础信息.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 10:
                    break
        return data
    
    @staticmethod
    def get_school_ranking(school_name=None, ranking_type='软科'):
        """获取院校排名"""
        file_map = {
            '软科': '2022排名_软科.xlsx',
            '校友会': '2022排名_校友会.xlsx',
            'QS': '2022排名_QS.xlsx',
            'U.S.News': '2022排名_U.S.News.xlsx',
            '泰晤士': '2022排名_泰晤士.xlsx'
        }
        if ranking_type not in file_map:
            return []
        file_path = os.path.join(DATA_DIR, f'5_参考数据/{file_map[ranking_type]}')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 15:
                    break
        return data
    
    @staticmethod
    def get_discipline_evaluation(school_name=None, discipline=None):
        """获取学科评估结果"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/第四轮学科评估.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[-1]):
                    continue
                if discipline and discipline not in str(row[2]):
                    continue
                data.append(row)
                if len(data) >= 15:
                    break
        return data


class MajorInfoSkill:
    @staticmethod
    def get_major_intro(major_name=None):
        """获取专业介绍"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业介绍及薪酬表.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None and row[4] is not None:
                if major_name and major_name not in str(row[4]):
                    continue
                data.append(row)
                if len(data) >= 5:
                    break
        return data
    
    @staticmethod
    def get_major_ranking(major_name=None):
        """获取专业排名"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业排名信息.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if major_name and major_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 15:
                    break
        return data
    
    @staticmethod
    def get_major_employment(major_name=None):
        """获取专业就业信息"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业就业信息.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if major_name and major_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 5:
                    break
        return data
    
    @staticmethod
    def get_major_satisfaction(major_name=None):
        """获取专业满意度"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业满意度.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if major_name and major_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 5:
                    break
        return data


class ReferenceDataSkill:
    @staticmethod
    def get_major_catalog():
        """获取本科专业目录"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/本科专业目录-2023.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))[:30]
    
    @staticmethod
    def get_b保研_info(school_name=None):
        """获取保研资格院校信息"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/保研资格院校保研率统计-2021.xls')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 10:
                    break
        return data
    
    @staticmethod
    def get_school_contact(school_name=None):
        """获取院校联系方式"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/高考院校电话.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 10:
                    break
        return data


SKILLS_MAP = {
    'get_batch_lines': ScoreLineSkill.get_batch_lines,
    'get_rank_by_score': ScoreLineSkill.get_rank_by_score,
    'get_school_scores': ScoreLineSkill.get_school_scores,
    'get_major_scores': ScoreLineSkill.get_major_scores,
    'get_enrollment_plan': EnrollmentPlanSkill.get_plan,
    'get_plan_by_year': EnrollmentPlanSkill.get_plan_by_year,
    'get_school_info': SchoolInfoSkill.get_school_info,
    'get_school_ranking': SchoolInfoSkill.get_school_ranking,
    'get_discipline_evaluation': SchoolInfoSkill.get_discipline_evaluation,
    'get_major_intro': MajorInfoSkill.get_major_intro,
    'get_major_ranking': MajorInfoSkill.get_major_ranking,
    'get_major_employment': MajorInfoSkill.get_major_employment,
    'get_major_satisfaction': MajorInfoSkill.get_major_satisfaction,
    'get_major_catalog': ReferenceDataSkill.get_major_catalog,
    'get_b保研_info': ReferenceDataSkill.get_b保研_info,
    'get_school_contact': ReferenceDataSkill.get_school_contact,
}

SKILL_DESCRIPTIONS = {
    'get_batch_lines': '获取高考批次线数据，可按年份和类别筛选',
    'get_rank_by_score': '根据分数查询全省排名',
    'get_school_scores': '获取院校录取分数线，可按院校名称、年份、最低分筛选',
    'get_major_scores': '获取专业录取分数线，可按院校名称、专业名称、年份筛选',
    'get_enrollment_plan': '获取招生计划数据，可按院校名称、专业名称筛选',
    'get_plan_by_year': '按年份获取招生计划',
    'get_school_info': '获取院校基础信息（地址、电话等）',
    'get_school_ranking': '获取院校排名（软科、校友会、QS等）',
    'get_discipline_evaluation': '获取学科评估结果',
    'get_major_intro': '获取专业介绍和培养方向',
    'get_major_ranking': '获取专业排名',
    'get_major_employment': '获取专业就业信息',
    'get_major_satisfaction': '获取专业满意度数据',
    'get_major_catalog': '获取本科专业目录',
    'get_b保研_info': '获取保研资格院校信息',
    'get_school_contact': '获取院校联系方式'
}

def call_skill(skill_name, **kwargs):
    """调用指定的skill"""
    if skill_name not in SKILLS_MAP:
        return f"错误：未知的skill: {skill_name}"
    
    print(f"🔍 正在调用技能: {skill_name} ({SKILL_DESCRIPTIONS[skill_name]})")
    print(f"   参数: {kwargs}")
    
    try:
        result = SKILLS_MAP[skill_name](**kwargs)
        print(f"✅ 技能调用成功，返回 {len(result)} 条数据")
        return result
    except Exception as e:
        print(f"❌ 技能调用失败: {e}")
        return f"错误: {e}"
