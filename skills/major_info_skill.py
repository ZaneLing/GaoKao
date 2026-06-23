import openpyxl
import os

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class MajorInfoSkill:
    @staticmethod
    def get_major_intro(major_name=None):
        """获取专业介绍"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业介绍及薪酬表.xlsx')
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            
            data = []
            for row in ws.iter_rows(values_only=True):
                if row[0] is not None and row[4] is not None:
                    if major_name and major_name not in str(row[4]):
                        continue
                    data.append(row)
                    if len(data) >= 5:
                        break
            
            return data
        except Exception:
            return []
    
    @staticmethod
    def get_major_basic_info(major_name=None):
        """获取专业基本信息（核心课程、就业去向等）"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业基本介绍.xlsx')
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            
            headers = next(ws.iter_rows(values_only=True))
            
            for row in ws.iter_rows(values_only=True):
                if row[0] is not None:
                    if major_name and major_name not in str(row[3]):
                        continue
                    
                    info = {}
                    for i, header in enumerate(headers):
                        if i < len(row):
                            info[str(header)] = row[i]
                    
                    return {
                        '专业名称': info.get('专业名称'),
                        '学科门类': info.get('学科门类'),
                        '专业类': info.get('专业类'),
                        '修业年限': info.get('修业年限'),
                        '授予学位': info.get('授予学位'),
                        '选考建议': info.get('选考（学科）建议'),
                        '就业率': info.get('就业率'),
                        '专业是什么': str(info.get('专业是什么'))[:200] if info.get('专业是什么') else None,
                        '专业学什么': str(info.get('专业学什么'))[:300] if info.get('专业学什么') else None,
                        '专业干什么': str(info.get('专业干什么'))[:200] if info.get('专业干什么') else None,
                        '就业去向': str(info.get('就业去向'))[:300] if info.get('就业去向') else None,
                        '就业地区分布': info.get('就业地区分布'),
                        '就业行业分布': info.get('就业行业分布'),
                        '就业岗位分布': info.get('就业岗位分布')
                    }
        except Exception:
            pass
        
        return None
    
    @staticmethod
    def get_major_ranking(major_name=None):
        """获取专业排名"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业排名信息.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if major_name and major_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 15:
                    break
        
        return data
    
    @staticmethod
    def analyze_major_ranking(major_name):
        """分析专业排名情况"""
        raw_data = MajorInfoSkill.get_major_ranking(major_name=major_name)
        
        if not raw_data:
            return {'major': major_name, 'error': '未找到专业排名数据'}
        
        rankings = []
        for row in raw_data:
            if len(row) >= 4:
                school = str(row[2]) if len(row) > 2 else '未知'
                rank = row[1] if isinstance(row[1], (int, float)) else 'N/A'
                level = str(row[3]) if len(row) > 3 else 'N/A'
                rankings.append({
                    'school': school,
                    'rank': rank,
                    'level': level
                })
        
        top_schools = [r['school'] for r in rankings[:5]]
        a_level_count = sum(1 for r in rankings if 'A' in r['level'])
        
        return {
            'major': major_name,
            'top_schools': top_schools,
            'total_ranked_schools': len(rankings),
            'A_level_schools': a_level_count,
            'ranking_details': rankings[:10]
        }
    
    @staticmethod
    def get_major_employment(major_name=None):
        """获取专业就业信息"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业就业信息.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if major_name and major_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 5:
                    break
        
        return data
    
    @staticmethod
    def get_major_satisfaction(major_name=None):
        """获取专业满意度"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业满意度.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if major_name and major_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 5:
                    break
        
        return data
    
    @staticmethod
    def analyze_major_employment_trend(major_name):
        """分析专业就业趋势"""
        employment_data = MajorInfoSkill.get_major_employment(major_name=major_name)
        
        if not employment_data:
            return {'major': major_name, 'error': '未找到就业数据'}
        
        return {
            'major': major_name,
            'employment_data': employment_data[:3],
            'suggestion': f"{major_name}专业就业信息已获取，可进一步分析就业前景"
        }
    
    @staticmethod
    def analyze_major_satisfaction(major_name):
        """分析专业满意度"""
        satisfaction_data = MajorInfoSkill.get_major_satisfaction(major_name=major_name)
        
        if not satisfaction_data:
            return {'major': major_name, 'error': '未找到满意度数据'}
        
        return {
            'major': major_name,
            'satisfaction_data': satisfaction_data[:3],
            'suggestion': f"{major_name}专业满意度数据已获取"
        }
    
    @staticmethod
    def compare_majors(majors):
        """对比多个专业"""
        comparison = {}
        
        for major in majors:
            ranking = MajorInfoSkill.analyze_major_ranking(major)
            employment = MajorInfoSkill.get_major_employment(major_name=major)
            satisfaction = MajorInfoSkill.get_major_satisfaction(major_name=major)
            
            comparison[major] = {
                'ranking': ranking,
                'employment': employment[:2],
                'satisfaction': satisfaction[:2]
            }
        
        return comparison
    
    @staticmethod
    def get_major_salary_info(major_name=None):
        """获取专业薪酬信息"""
        file_path = os.path.join(DATA_DIR, '4_专业信息/专业介绍及薪酬表.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None and row[4] is not None:
                if major_name and major_name not in str(row[4]):
                    continue
                
                salary_info = {}
                salary_info['专业名称'] = row[4]
                for i, val in enumerate(row):
                    if isinstance(val, (int, float)) and val > 1000:
                        salary_info[f'薪酬{i}'] = val
                
                if salary_info:
                    data.append(salary_info)
                    if len(data) >= 5:
                        break
        
        return data
    
    @staticmethod
    def analyze_major_popularity(major_name):
        """分析专业冷热程度"""
        basic_info = MajorInfoSkill.get_major_basic_info(major_name)
        ranking = MajorInfoSkill.analyze_major_ranking(major_name)
        
        popularity_score = 0
        factors = []
        
        if basic_info:
            employment_rate = basic_info.get('就业率')
            if employment_rate and isinstance(employment_rate, (int, float)) and employment_rate >= 90:
                popularity_score += 25
                factors.append('就业率高')
            elif employment_rate and isinstance(employment_rate, (int, float)) and employment_rate >= 80:
                popularity_score += 15
                factors.append('就业率较高')
        
        if ranking:
            a_count = ranking.get('A_level_schools', 0)
            total = ranking.get('total_ranked_schools', 0)
            if a_count >= 5:
                popularity_score += 30
                factors.append('顶尖院校多')
            elif a_count >= 2:
                popularity_score += 20
                factors.append('有顶尖院校')
            if total >= 20:
                popularity_score += 20
                factors.append('开设院校多')
            elif total >= 10:
                popularity_score += 10
                factors.append('开设院校适中')
        
        hot_keywords = ['计算机', '人工智能', '大数据', '临床医学', '口腔医学', '金融', '电子信息']
        warm_keywords = ['医学', '药学', '护理', '生物', '环境', '材料', '机械', '土木']
        
        if any(kw in major_name for kw in hot_keywords):
            popularity_score += 25
            factors.append('当前热门专业')
        elif any(kw in major_name for kw in warm_keywords):
            popularity_score += 10
            factors.append('需求稳定专业')
        
        if popularity_score >= 80:
            level = '热门'
        elif popularity_score >= 60:
            level = '较热'
        elif popularity_score >= 40:
            level = '适中'
        else:
            level = '较冷'
        
        return {
            'major': major_name,
            'popularity_score': popularity_score,
            'popularity_level': level,
            'factors': factors,
            'basic_info': basic_info,
            'ranking': ranking
        }
    
    @staticmethod
    def analyze_major_adjustment_risk(major_name, school_name=None):
        """分析专业调剂风险"""
        major_basic_info = MajorInfoSkill.get_major_basic_info(major_name)
        
        risk_score = 0
        factors = []
        
        employment_rate = major_basic_info.get('就业率') if major_basic_info else None
        if employment_rate and isinstance(employment_rate, (int, float)):
            if employment_rate >= 90:
                risk_score += 30
                factors.append('就业率高，竞争激烈')
            elif employment_rate >= 80:
                risk_score += 20
                factors.append('就业率较高')
        
        popularity = MajorInfoSkill.analyze_major_popularity(major_name)
        if popularity['popularity_level'] == '热门':
            risk_score += 30
            factors.append('热门专业，报考人数多')
        elif popularity['popularity_level'] == '较热':
            risk_score += 20
            factors.append('较热门专业')
        
        if risk_score >= 60:
            risk_level = '高'
        elif risk_score >= 40:
            risk_level = '中'
        else:
            risk_level = '低'
        
        return {
            'major': major_name,
            'school': school_name,
            'adjustment_risk_score': risk_score,
            'adjustment_risk_level': risk_level,
            'risk_factors': factors
        }
    
    @staticmethod
    def analyze_major_group_admission_details(major_name, school_name=None):
        """分析专业组录取详情（大类招生、分流规则等）"""
        basic_info = MajorInfoSkill.get_major_basic_info(major_name)
        
        has_broad_recruitment = False
        broad_info = {}
        
        if basic_info:
            major_desc = basic_info.get('专业是什么', '')
            if '大类' in str(major_desc) or '类' in str(major_desc):
                has_broad_recruitment = True
                broad_info['description'] = str(major_desc)[:200]
        
        return {
            'major': major_name,
            'school': school_name,
            'has_broad_recruitment': has_broad_recruitment,
            'broad_recruitment_info': broad_info,
            'basic_info': basic_info
        }
    
    @staticmethod
    def analyze_major_career_path(major_name):
        """分析专业就业方向和培养路径"""
        basic_info = MajorInfoSkill.get_major_basic_info(major_name)
        
        if not basic_info:
            return {'major': major_name, 'error': '未找到专业信息'}
        
        career_info = {
            'major': major_name,
            'degree': basic_info.get('授予学位'),
            'duration': basic_info.get('修业年限'),
            'employment_rate': basic_info.get('就业率'),
            'what_it_is': basic_info.get('专业是什么'),
            'what_to_study': basic_info.get('专业学什么'),
            'what_to_do': basic_info.get('专业干什么'),
            'job_destinations': basic_info.get('就业去向'),
            'regional_distribution': basic_info.get('就业地区分布'),
            'industry_distribution': basic_info.get('就业行业分布'),
            'position_distribution': basic_info.get('就业岗位分布')
        }
        
        return career_info
    
    @staticmethod
    def analyze_major_comprehensive(major_name):
        """综合分析专业（排名、就业、薪酬、满意度）"""
        ranking = MajorInfoSkill.analyze_major_ranking(major_name)
        employment = MajorInfoSkill.get_major_employment(major_name=major_name)
        satisfaction = MajorInfoSkill.get_major_satisfaction(major_name=major_name)
        salary = MajorInfoSkill.get_major_salary_info(major_name=major_name)
        basic_info = MajorInfoSkill.get_major_basic_info(major_name)
        popularity = MajorInfoSkill.analyze_major_popularity(major_name)
        career_path = MajorInfoSkill.analyze_major_career_path(major_name)
        
        score = 0
        factors = []
        
        if 'A_level_schools' in ranking and ranking['A_level_schools'] >= 3:
            score += 30
            factors.append('学科实力强')
        elif 'A_level_schools' in ranking and ranking['A_level_schools'] >= 1:
            score += 20
            factors.append('学科实力较强')
        
        if employment:
            score += 25
            factors.append('就业数据充足')
        
        if satisfaction:
            score += 15
            factors.append('满意度数据可查')
        
        if salary:
            score += 30
            factors.append('薪酬数据可查')
        
        return {
            'major': major_name,
            'overall_score': score,
            'factors': factors,
            'ranking': ranking,
            'employment': employment[:2],
            'satisfaction': satisfaction[:2],
            'salary': salary[:2],
            'basic_info': basic_info,
            'popularity': popularity,
            'career_path': career_path,
            'recommendation': '强烈推荐' if score >= 80 else '推荐' if score >= 60 else '谨慎选择'
        }

SKILLS_MAP = {
    'get_major_intro': MajorInfoSkill.get_major_intro,
    'get_major_basic_info': MajorInfoSkill.get_major_basic_info,
    'get_major_ranking': MajorInfoSkill.get_major_ranking,
    'analyze_major_ranking': MajorInfoSkill.analyze_major_ranking,
    'get_major_employment': MajorInfoSkill.get_major_employment,
    'get_major_satisfaction': MajorInfoSkill.get_major_satisfaction,
    'analyze_major_employment_trend': MajorInfoSkill.analyze_major_employment_trend,
    'analyze_major_satisfaction': MajorInfoSkill.analyze_major_satisfaction,
    'compare_majors': MajorInfoSkill.compare_majors,
    'get_major_salary_info': MajorInfoSkill.get_major_salary_info,
    'analyze_major_comprehensive': MajorInfoSkill.analyze_major_comprehensive,
    'analyze_major_popularity': MajorInfoSkill.analyze_major_popularity,
    'analyze_major_adjustment_risk': MajorInfoSkill.analyze_major_adjustment_risk,
    'analyze_major_group_admission_details': MajorInfoSkill.analyze_major_group_admission_details,
    'analyze_major_career_path': MajorInfoSkill.analyze_major_career_path,
}

SKILL_DESCRIPTIONS = {
    'get_major_intro': '获取专业介绍和培养方向',
    'get_major_basic_info': '获取专业基本信息（核心课程、就业去向等）',
    'get_major_ranking': '获取专业排名',
    'analyze_major_ranking': '分析专业排名情况（顶尖院校、A类数量）',
    'get_major_employment': '获取专业就业信息',
    'get_major_satisfaction': '获取专业满意度数据',
    'analyze_major_employment_trend': '分析专业就业趋势',
    'analyze_major_satisfaction': '分析专业满意度',
    'compare_majors': '对比多个专业的综合情况',
    'get_major_salary_info': '获取专业薪酬信息',
    'analyze_major_comprehensive': '综合分析专业（排名、就业、薪酬、满意度）',
    'analyze_major_popularity': '分析专业冷热程度',
    'analyze_major_adjustment_risk': '分析专业调剂风险',
    'analyze_major_group_admission_details': '分析专业组录取详情（大类招生、分流规则等）',
    'analyze_major_career_path': '分析专业就业方向和培养路径',
}
