import openpyxl
import xlrd
import os
from .data_utils import clean_text, read_rows

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class ReferenceDataSkill:
    @staticmethod
    def get_major_catalog():
        """获取本科专业目录"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/本科专业目录-2023.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        return list(ws.iter_rows(values_only=True))[:30]
    
    @staticmethod
    def search_major_catalog(keyword):
        """搜索本科专业目录"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/本科专业目录-2023.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        results = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None and keyword in str(row):
                results.append(row)
                if len(results) >= 10:
                    break
        
        return results
    
    @staticmethod
    def get_b保研_info(school_name=None):
        """获取保研资格院校信息"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/保研资格院校保研率统计-2021.xls')
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        
        data = []
        for i in range(min(ws.nrows, 50)):
            row = ws.row_values(i)
            if row[0] != '':
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(tuple(row))
                if len(data) >= 10:
                    break
        
        return data
    
    @staticmethod
    def analyze_b保研_trend():
        """分析保研率趋势"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/保研资格院校保研率统计-2021.xls')
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        
        rates = []
        for i in range(min(ws.nrows, 100)):
            row = ws.row_values(i)
            if row[0] != '' and len(row) > 1:
                rate = row[1]
                if isinstance(rate, (int, float)) and rate > 0 and rate < 100:
                    rates.append(rate)
        
        if rates:
            avg_rate = sum(rates) / len(rates)
            high_rate_schools = sum(1 for r in rates if r >= 20)
            return {
                'total_schools': len(rates),
                'avg_rate': round(avg_rate, 1),
                'high_rate_schools': high_rate_schools,
                'suggestion': f"全国平均保研率约{round(avg_rate, 1)}%，保研率20%以上的高水平院校有{high_rate_schools}所"
            }
        
        return {'error': '无法提取保研率数据'}
    
    @staticmethod
    def get_school_contact(school_name=None):
        """获取院校联系方式"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/高考院校电话.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 10:
                    break
        
        return data
    
    @staticmethod
    def get_province_schools(province=None):
        """按省份获取院校列表"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/高考院校电话.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if province and province not in str(row):
                    continue
                data.append(row)
                if len(data) >= 20:
                    break
        
        return data
    
    @staticmethod
    def get_discipline_evaluation_summary():
        """获取学科评估汇总"""
        level_counts = {'A+': 0, 'A': 0, 'A-': 0, 'B+': 0, 'B': 0, 'B-': 0, 'C+': 0, 'C': 0, 'C-': 0}

        for row in read_rows('5_参考数据/第四轮学科评估.xlsx'):
            level = clean_text(row.get('评估结果'))
            if level in level_counts:
                level_counts[level] += 1
        
        total = sum(level_counts.values())
        return {
            'total_disciplines': total,
            'level_counts': level_counts,
            'summary': f"第四轮学科评估共覆盖{total}个学科，其中A+等级{level_counts['A+']}个，A等级{level_counts['A']}个，A-等级{level_counts['A-']}个"
        }
    
    @staticmethod
    def get_dual_first_class_info():
        """获取双一流信息"""
        return {
            'description': '双一流建设是中国高等教育领域继"211工程"、"985工程"之后的又一国家战略',
            'categories': {
                '世界一流大学': {
                    'A类': ['北京大学', '清华大学', '复旦大学', '上海交通大学', '浙江大学', '南京大学', '中国人民大学', '中国科学技术大学'],
                    'B类': ['东北大学', '郑州大学', '湖南大学']
                },
                '世界一流学科': {
                    '安徽': ['安徽大学', '合肥工业大学', '中国科学技术大学']
                }
            },
            'suggestion': '双一流院校在资源投入、学科建设等方面有明显优势，建议优先考虑'
        }
    
    @staticmethod
    def get_211_985_info():
        """获取211/985信息"""
        return {
            'description': '211工程和985工程是中国高等教育的重点建设项目',
            'stats': {
                '985工程': 39,
                '211工程': 112,
                '211包含985': True
            },
            'regional_info': {
                '安徽': {
                    '985': ['中国科学技术大学'],
                    '211': ['中国科学技术大学', '安徽大学', '合肥工业大学']
                }
            },
            'suggestion': '985/211院校在就业、升学等方面有优势，但也要考虑专业实力和个人发展方向'
        }

SKILLS_MAP = {
    'get_major_catalog': ReferenceDataSkill.get_major_catalog,
    'search_major_catalog': ReferenceDataSkill.search_major_catalog,
    'get_b保研_info': ReferenceDataSkill.get_b保研_info,
    'analyze_b保研_trend': ReferenceDataSkill.analyze_b保研_trend,
    'get_school_contact': ReferenceDataSkill.get_school_contact,
    'get_province_schools': ReferenceDataSkill.get_province_schools,
    'get_discipline_evaluation_summary': ReferenceDataSkill.get_discipline_evaluation_summary,
    'get_dual_first_class_info': ReferenceDataSkill.get_dual_first_class_info,
    'get_211_985_info': ReferenceDataSkill.get_211_985_info,
}

SKILL_DESCRIPTIONS = {
    'get_major_catalog': '获取本科专业目录',
    'search_major_catalog': '搜索本科专业目录',
    'get_b保研_info': '获取保研资格院校信息',
    'analyze_b保研_trend': '分析保研率趋势',
    'get_school_contact': '获取院校联系方式',
    'get_province_schools': '按省份获取院校列表',
    'get_discipline_evaluation_summary': '获取学科评估汇总',
    'get_dual_first_class_info': '获取双一流建设高校信息',
    'get_211_985_info': '获取211/985工程高校信息',
}
