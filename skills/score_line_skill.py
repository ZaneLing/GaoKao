import openpyxl
import os

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class ScoreLineSkill:
    @staticmethod
    def get_batch_lines_trend(years=None, category=None, province='安徽', subject=None, batch=None):
        """获取多年份批次线趋势数据"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/批次线-2021-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        if years is None:
            years = [2021, 2022, 2023, 2024]
        
        # 处理subject参数映射
        if subject:
            subject_map = {'物理类': '物理', '历史类': '历史', '理科': '物理', '文科': '历史'}
            category = subject_map.get(subject, subject)
        
        trend_data = {}
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None and isinstance(row[1], int) and row[1] in years:
                if category and category != row[3]:
                    continue
                if batch and batch not in str(row[2]):
                    continue
                year = row[1]
                if year not in trend_data:
                    trend_data[year] = []
                trend_data[year].append({
                    '批次': row[2],
                    '科类': row[3],
                    '分数线': row[5],
                    '位次': row[6]
                })
        
        return trend_data
    
    @staticmethod
    def analyze_batch_trend(category='理科', province='安徽', subject=None, batch=None):
        """分析批次线多年趋势"""
        # 处理subject参数
        if subject:
            subject_map = {'物理类': '物理', '历史类': '历史', '理科': '物理', '文科': '历史'}
            category = subject_map.get(subject, subject)
        
        trend_data = ScoreLineSkill.get_batch_lines_trend(category=category, batch=batch)
        analysis = []
        
        for year in sorted(trend_data.keys()):
            for item in trend_data[year]:
                analysis.append([year, item['批次'], item['科类'], item['分数线'], item['位次']])
        
        return analysis
    
    @staticmethod
    def get_rank_by_score(score, year=2024, subject=None):
        """根据分数查询全省排名"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/一分一段-2017-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        for row in ws.iter_rows(values_only=True):
            if row[0] is None:
                continue
            if isinstance(row[8], (int, float)) and int(row[8]) == score:
                row_subject = str(row[2]) if row[2] else ''
                if subject:
                    subject_map = {'物理类': '物理', '历史类': '历史', '理科': '物理', '文科': '历史'}
                    target_subject = subject_map.get(subject, subject)
                    if target_subject not in row_subject:
                        continue
                return {
                    '分数': row[8],
                    '累计人数': row[4] if row[4] else row[5],
                    '位次': row[5],
                    '年份': row[1],
                    '科目': row[2]
                }
        
        return None
    
    @staticmethod
    def get_score_by_rank(rank, year=2024):
        """根据排名查询对应分数"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/一分一段-2017-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        for row in ws.iter_rows(values_only=True):
            if isinstance(row[2], (int, float)) and int(row[2]) <= rank:
                return {
                    '分数': row[0],
                    '累计人数': row[1],
                    '位次': row[2],
                    '年份': year
                }
        
        return None
    
    @staticmethod
    def get_school_scores_trend(school_name=None, years=None):
        """获取院校多年份分数线趋势"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/院校分数-2020-2024.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        if years is None:
            years = [2020, 2021, 2022, 2023, 2024]
        
        trend_data = []
        for row in ws.iter_rows(values_only=True):
            if row[1] is not None:
                if school_name and school_name not in str(row[1]):
                    continue
                trend_data.append(row)
        
        return trend_data[:30]
    
    @staticmethod
    def analyze_school_score_trend(school_name):
        """分析院校分数线趋势"""
        raw_data = ScoreLineSkill.get_school_scores_trend(school_name=school_name)
        
        if not raw_data:
            return {'school': school_name, 'error': '未找到院校数据'}
        
        scores = []
        for row in raw_data:
            for i, val in enumerate(row):
                if isinstance(val, (int, float)) and val > 400:
                    scores.append(val)
        
        if scores:
            avg_score = sum(scores) / len(scores)
            min_score = min(scores)
            max_score = max(scores)
            volatility = (max_score - min_score) / avg_score * 100
            
            return {
                'school': school_name,
                'years': len(set([str(row[1]) for row in raw_data if len(row) > 1])),
                'avg_score': round(avg_score, 1),
                'min_score': min_score,
                'max_score': max_score,
                'volatility': round(volatility, 1),
                'trend': '上升' if scores[-1] > scores[0] else '下降' if scores[-1] < scores[0] else '稳定',
                'samples': raw_data[:5]
            }
        return {'school': school_name, 'error': '无法提取分数数据'}
    
    @staticmethod
    def get_major_scores(major_name=None, school_name=None, years=None):
        """获取专业分数线"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/专业分数-2024-考试院.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        if years is None:
            years = [2024]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                if major_name and major_name not in str(row):
                    continue
                data.append(row)
                if len(data) >= 20:
                    break
        
        return data
    
    @staticmethod
    def analyze_major_score_comparison(major_name, schools=None):
        """对比分析多所院校的同一专业分数线"""
        data = ScoreLineSkill.get_major_scores(major_name=major_name)
        
        comparison = []
        for row in data:
            school = str(row[0]) if len(row) > 0 else ''
            if schools and school not in schools:
                continue
            
            scores = []
            for i in range(1, len(row)):
                if isinstance(row[i], (int, float)) and row[i] > 0:
                    scores.append(row[i])
            
            if scores:
                comparison.append({
                    'school': school,
                    'major': major_name,
                    'avg_score': round(sum(scores) / len(scores), 1),
                    'min_score': min(scores),
                    'max_score': max(scores),
                    'data': row
                })
        
        comparison.sort(key=lambda x: x['avg_score'], reverse=True)
        return comparison[:10]
    
    @staticmethod
    def predict_admission_score(school_name=None, user_score=None, score=None, rank=None, subject=None):
        """预测录取可能性"""
        # 处理参数别名
        if score and not user_score:
            user_score = score
        
        if not school_name:
            return {'error': '请提供学校名称'}
        
        if not user_score:
            return {'error': '请提供分数'}
        
        analysis = ScoreLineSkill.analyze_school_score_trend(school_name)
        
        if 'error' in analysis:
            return {'school': school_name, 'error': analysis['error']}
        
        avg_score = analysis['avg_score']
        min_score = analysis['min_score']
        max_score = analysis['max_score']
        
        if user_score >= max_score:
            probability = '很高 (90%以上)'
            suggestion = '可以作为稳或保的选择'
        elif user_score >= avg_score:
            probability = '较高 (70%-90%)'
            suggestion = '可以作为稳的选择'
        elif user_score >= min_score:
            probability = '一般 (40%-70%)'
            suggestion = '需要谨慎考虑，可作为冲的选择'
        elif user_score >= min_score - 10:
            probability = '较低 (10%-40%)'
            suggestion = '冲刺选择，风险较大'
        else:
            probability = '很低 (10%以下)'
            suggestion = '不建议报考'
        
        return {
            'school': school_name,
            'user_score': user_score,
            'user_rank': rank,
            'subject': subject,
            'school_avg_score': avg_score,
            'school_min_score': min_score,
            'school_max_score': max_score,
            'probability': probability,
            'suggestion': suggestion
        }

SKILLS_MAP = {
    'get_batch_lines_trend': ScoreLineSkill.get_batch_lines_trend,
    'analyze_batch_trend': ScoreLineSkill.analyze_batch_trend,
    'get_rank_by_score': ScoreLineSkill.get_rank_by_score,
    'get_score_by_rank': ScoreLineSkill.get_score_by_rank,
    'get_school_scores_trend': ScoreLineSkill.get_school_scores_trend,
    'analyze_school_score_trend': ScoreLineSkill.analyze_school_score_trend,
    'get_major_scores': ScoreLineSkill.get_major_scores,
    'analyze_major_score_comparison': ScoreLineSkill.analyze_major_score_comparison,
    'predict_admission_score': ScoreLineSkill.predict_admission_score,
}

SKILL_DESCRIPTIONS = {
    'get_batch_lines_trend': '获取多年份批次线趋势数据(参数: years, category, province, subject, batch)',
    'analyze_batch_trend': '分析批次线多年变化趋势(参数: category, province, subject, batch)',
    'get_rank_by_score': '根据分数查询全省排名(参数: score, year, subject)',
    'get_score_by_rank': '根据排名查询对应分数(参数: rank, year)',
    'get_school_scores_trend': '获取院校多年份分数线数据(参数: school_name, years)',
    'analyze_school_score_trend': '分析院校分数线趋势（平均分、波动、走向）(参数: school_name)',
    'get_major_scores': '获取专业分数线(参数: major_name, year)',
    'analyze_major_score_comparison': '对比分析多所院校的同一专业分数线(参数: major_name, school_names)',
    'predict_admission_score': '根据用户分数预测录取可能性(参数: school_name, user_score, score, rank, subject)',
}
