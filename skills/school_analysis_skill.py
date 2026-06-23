"""
学校综合信息分析技能
从多个数据源提取并整合学校的完整信息
"""

import os
import openpyxl
from datetime import datetime

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class SchoolAnalysisSkill:
    """学校综合信息分析"""
    
    @staticmethod
    def _load_score_data(year=2024):
        """加载专业组分数线数据"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/专业组分数-2024.xlsx')
        if not os.path.exists(file_path):
            return []
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        return data
    
    @staticmethod
    def _load_plan_data(year=2025):
        """加载招生计划数据"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/招生计划/招生计划-2025.xlsx')
        if not os.path.exists(file_path):
            return []
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        return data
    
    @staticmethod
    def _load_major_scores(year=2024):
        """加载专业分数线数据"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/专业分数-2024-考试院.xlsx')
        if not os.path.exists(file_path):
            return []
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        return data
    
    @staticmethod
    def get_comprehensive_school_analysis(school_name):
        """
        获取学校综合信息分析
        整合历年分数线、专业组、招生计划、专业详情等
        """
        if not school_name:
            return {"error": "请指定学校名称"}
        
        result = {
            "学校名称": school_name,
            "分析时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "基础信息": {},
            "历年分数线": [],
            "专业组信息": [],
            "招生计划": [],
            "专业详情": [],
            "录取分析": {}
        }
        
        # 1. 获取基础信息
        result["基础信息"] = SchoolAnalysisSkill._get_school_basic_info(school_name)
        
        # 2. 获取历年分数线
        result["历年分数线"] = SchoolAnalysisSkill._get_school_score_history(school_name)
        
        # 3. 获取专业组信息
        result["专业组信息"] = SchoolAnalysisSkill._get_school_major_groups(school_name)
        
        # 4. 获取招生计划
        result["招生计划"] = SchoolAnalysisSkill._get_school_enrollment_plan(school_name)
        
        # 5. 获取专业详情
        result["专业详情"] = SchoolAnalysisSkill._get_school_majors(school_name)
        
        # 6. 生成录取分析
        result["录取分析"] = SchoolAnalysisSkill._analyze_admission(school_name, result)
        
        return result
    
    @staticmethod
    def _get_school_basic_info(school_name):
        """获取学校基础信息"""
        file_path = os.path.join(DATA_DIR, '3_院校信息/院校基础信息.xlsx')
        if not os.path.exists(file_path):
            return {}
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = None
        
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = row
                continue
            if row[0] and school_name in str(row[0]):
                info = {}
                for i, h in enumerate(headers):
                    if i < len(row) and row[i]:
                        info[str(h)] = row[i]
                return {
                    "学校名称": info.get("学校名称"),
                    "所在省": info.get("所在省"),
                    "城市": info.get("城市"),
                    "院校类型": info.get("类型"),
                    "是否985": info.get("是否985", "否"),
                    "是否211": info.get("是否211", "否"),
                    "是否双一流": info.get("是否双一流", "否"),
                    "是否重点": info.get("国重/省重", ""),
                    "公私性质": info.get("公私性质"),
                    "本科/专科": info.get("本科/专科"),
                    "保研率": info.get("保研率", "N/A"),
                    "国家特色专业": info.get("国家特色专业", ""),
                    "省特色专业": info.get("省特色专业", ""),
                    "硕士点": info.get("硕士点（个）", "N/A"),
                    "博士点": info.get("博士点（个）", "N/A"),
                    "学科评估": info.get("评估结果", ""),
                    "招办电话": info.get("招办电话", ""),
                    "官网": info.get("官网", "")
                }
        return {}
    
    @staticmethod
    def _get_school_score_history(school_name):
        """获取学校历年分数线"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/院校分数-2020-2024.xlsx')
        if not os.path.exists(file_path):
            return []
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = None
        results = []
        
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h) if h else f"列{i}" for i, h in enumerate(row)]
                continue  # 跳过标题行
            # 第2列是学校名称
            if row[1] and school_name in str(row[1]):
                record = {}
                for i, h in enumerate(headers):
                    if i < len(row):
                        record[h] = row[i]
                results.append(record)
        
        return results
    
    @staticmethod
    def _get_school_major_groups(school_name):
        """获取学校专业组信息（暂无数据源）"""
        # 注意：专业组数据在其他文件中，这里返回历年分数线作为参考
        return SchoolAnalysisSkill._get_school_score_history(school_name)
    
    @staticmethod
    def _get_school_enrollment_plan(school_name):
        """获取学校招生计划"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/招生计划/招生计划-2025.xlsx')
        if not os.path.exists(file_path):
            return []
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = None
        results = []
        
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h) if h else f"列{i}" for i, h in enumerate(row)]
                continue  # 跳过标题行
            if row[1] and school_name in str(row[1]):  # 第2列是学校名称
                record = {}
                for i, h in enumerate(headers):
                    if i < len(row):
                        record[h] = row[i]
                results.append(record)
        
        return results
    
    @staticmethod
    def _get_school_majors(school_name):
        """获取学校专业详情"""
        file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/专业分数-2024-考试院.xlsx')
        if not os.path.exists(file_path):
            return []
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        headers = None
        results = []
        
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(h) if h else f"列{i}" for i, h in enumerate(row)]
                continue  # 跳过标题行
            # 第2列是院校名称
            if row[1] and school_name in str(row[1]):
                record = {}
                for i, h in enumerate(headers):
                    if i < len(row):
                        record[h] = row[i]
                results.append(record)
                if len(results) >= 50:  # 限制数量
                    break
        
        return results
    
    @staticmethod
    def _analyze_admission(school_name, data):
        """生成录取分析"""
        analysis = {
            "总专业组数": len(data.get("专业组信息", [])),
            "总招生专业数": len(data.get("专业详情", [])),
            "近3年分数线": [],
            "招生特点": [],
            "报考建议": []
        }
        
        # 提取近3年分数线
        score_history = data.get("历年分数线", [])
        for record in score_history[:3]:
            year = record.get("年份", record.get("年", "N/A"))
            score = record.get("分数线", record.get("最低分", record.get("录取分数", "N/A")))
            rank = record.get("最低位次", record.get("位次", "N/A"))
            if score and score != "N/A":
                analysis["近3年分数线"].append({
                    "年份": year,
                    "分数线": score,
                    "最低位次": rank
                })
        
        # 分析招生特点
        plan_count = len(data.get("招生计划", []))
        major_count = len(data.get("专业详情", []))
        
        if plan_count > 0:
            analysis["招生特点"].append(f"2025年招生{plan_count}个专业组/专业")
        
        basic_info = data.get("基础信息", {})
        if basic_info.get("是否985") == "是":
            analysis["招生特点"].append("985高校，层次较高，竞争激烈")
        if basic_info.get("是否211") == "是":
            analysis["招生特点"].append("211高校，实力雄厚")
        if basic_info.get("是否双一流") == "是":
            analysis["招生特点"].append("双一流建设高校")
        
        # 报考建议
        if basic_info.get("保研率") and basic_info.get("保研率") != "N/A":
            try:
                baoyan_rate = float(basic_info.get("保研率", 0))
                if baoyan_rate > 20:
                    analysis["报考建议"].append(f"保研率较高({baoyan_rate}%)，适合有读研意向的考生")
                elif baoyan_rate > 10:
                    analysis["报考建议"].append(f"保研率中等({baoyan_rate}%)，有一定保研机会")
            except:
                pass
        
        if major_count > 30:
            analysis["报考建议"].append("招生专业丰富，选择余地大")
        elif major_count > 10:
            analysis["报考建议"].append("招生专业适中，建议重点关注优势专业")
        else:
            analysis["报考建议"].append("招生专业较少，需仔细研究具体专业")
        
        return analysis
    
    @staticmethod
    def format_comprehensive_report(data):
        """格式化综合报告为可读文本"""
        if "error" in data:
            return f"错误: {data['error']}"
        
        lines = []
        lines.append("="*80)
        lines.append(f"🏫 {data['学校名称']} - 综合信息分析报告")
        lines.append("="*80)
        lines.append(f"📅 分析时间: {data['分析时间']}")
        lines.append("")
        
        # 基础信息
        basic = data.get("基础信息", {})
        if basic:
            lines.append("-"*80)
            lines.append("📋 【基础信息】")
            lines.append("-"*80)
            lines.append(f"  院校类型: {basic.get('院校类型', 'N/A')}")
            lines.append(f"  所在地区: {basic.get('所在省', 'N/A')} {basic.get('城市', '')}")
            levels = [str(basic.get('是否985', '')), str(basic.get('是否211', '')), str(basic.get('是否双一流', ''))]
            level_str = ' '.join(filter(None, levels)) or '普通本科'
            lines.append(f"  院校层次: {level_str}")
            lines.append(f"  公私性质: {basic.get('公私性质', 'N/A')}")
            lines.append(f"  本科/专科: {basic.get('本科/专科', 'N/A')}")
            lines.append(f"  保研率: {basic.get('保研率', 'N/A')}")
            lines.append(f"  硕士点: {basic.get('硕士点', 'N/A')}个")
            lines.append(f"  博士点: {basic.get('博士点', 'N/A')}个")
            lines.append(f"  学科评估: {basic.get('学科评估', 'N/A')}")
            lines.append(f"  国家特色专业: {basic.get('国家特色专业', 'N/A') or '无'}")
            lines.append(f"  省特色专业: {basic.get('省特色专业', 'N/A') or '无'}")
            lines.append(f"  招办电话: {basic.get('招办电话', 'N/A')}")
            lines.append(f"  官网: {basic.get('官网', 'N/A')}")
            lines.append("")
        
        # 历年分数线
        score_history = data.get("历年分数线", [])
        if score_history:
            lines.append("-"*80)
            lines.append("📈 【历年分数线】")
            lines.append("-"*80)
            lines.append(f"{'年份':<8} {'科类':<10} {'批次':<10} {'分数线':<10} {'最低位次':<15}")
            lines.append("-"*60)
            for record in score_history[:10]:
                year = record.get('年份', record.get('年', 'N/A'))
                category = record.get('科类', 'N/A')
                batch = record.get('批次', record.get('录取批次', 'N/A'))
                score = record.get('分数线', record.get('最低分', record.get('录取分数', 'N/A')))
                rank = record.get('最低位次', record.get('位次', 'N/A'))
                lines.append(f"{str(year):<8} {str(category):<10} {str(batch):<10} {str(score):<10} {str(rank):<15}")
            lines.append("")
        
        # 专业组信息
        major_groups = data.get("专业组信息", [])
        if major_groups:
            lines.append("-"*80)
            lines.append("🎯 【专业组信息】（2024年）")
            lines.append("-"*80)
            lines.append(f"{'专业组代码':<12} {'科目要求':<10} {'最高分':<10} {'最低分':<10} {'平均分':<10} {'位次':<15}")
            lines.append("-"*80)
            for group in major_groups[:20]:
                group_code = group.get('专业组代码', group.get('组号', group.get('组', 'N/A')))
                subject = group.get('科目要求', group.get('选科要求', 'N/A'))
                high = group.get('最高分', 'N/A')
                low = group.get('最低分', 'N/A')
                avg = group.get('平均分', 'N/A')
                rank = group.get('最低位次', group.get('位次', 'N/A'))
                lines.append(f"{str(group_code):<12} {str(subject):<10} {str(high):<10} {str(low):<10} {str(avg):<10} {str(rank):<15}")
            lines.append("")
        
        # 招生计划
        enrollment = data.get("招生计划", [])
        if enrollment:
            lines.append("-"*80)
            lines.append("📊 【招生计划】（2025年）")
            lines.append("-"*80)
            lines.append(f"{'专业组':<12} {'包含专业':<40} {'计划数':<10} {'科目要求':<10}")
            lines.append("-"*80)
            for plan in enrollment[:20]:
                group = plan.get('专业组代码', plan.get('组号', plan.get('组', 'N/A')))
                major = str(plan.get('包含专业', plan.get('专业', 'N/A')))[:38]
                plan_num = plan.get('计划总数', plan.get('计划数', plan.get('招生人数', 'N/A')))
                subject_req = plan.get('科目要求', plan.get('选科要求', 'N/A'))
                lines.append(f"{str(group):<12} {major:<40} {str(plan_num):<10} {str(subject_req):<10}")
            lines.append("")
        
        # 专业详情
        majors = data.get("专业详情", [])
        if majors:
            lines.append("-"*80)
            lines.append("📚 【专业详情】（2024年）")
            lines.append("-"*80)
            lines.append(f"{'专业名称':<20} {'最高分':<10} {'最低分':<10} {'平均分':<10} {'位次':<15}")
            lines.append("-"*80)
            for major in majors[:30]:
                name = str(major.get('专业名称', major.get('专业', 'N/A')))[:18]
                high = major.get('最高分', 'N/A')
                low = major.get('最低分', 'N/A')
                avg = major.get('平均分', 'N/A')
                rank = major.get('最低位次', major.get('位次', 'N/A'))
                lines.append(f"{name:<20} {str(high):<10} {str(low):<10} {str(avg):<10} {str(rank):<15}")
            lines.append("")
        
        # 录取分析
        analysis = data.get("录取分析", {})
        if analysis:
            lines.append("-"*80)
            lines.append("💡 【录取分析与建议】")
            lines.append("-"*80)
            
            if analysis.get("近3年分数线"):
                lines.append("近3年分数线:")
                for score_info in analysis["近3年分数线"]:
                    lines.append(f"  {score_info['年份']}年: {score_info['分数线']}分 (位次: {score_info['最低位次']})")
            
            if analysis.get("招生特点"):
                lines.append("\n招生特点:")
                for feature in analysis["招生特点"]:
                    lines.append(f"  • {feature}")
            
            if analysis.get("报考建议"):
                lines.append("\n报考建议:")
                for suggestion in analysis["报考建议"]:
                    lines.append(f"  ✓ {suggestion}")
            
            lines.append("")
        
        lines.append("="*80)
        lines.append("📝 报告说明:")
        lines.append("  1. 数据来源于安徽省教育考试院官方数据")
        lines.append("  2. 分数线和位次为历史数据，仅供参考")
        lines.append("  3. 招生计划为2025年数据，最终以官方公布为准")
        lines.append("  4. 建议结合自身排名和分数综合考虑")
        lines.append("="*80)
        
        return "\n".join(lines)


# Skill注册
SKILLS_MAP = {
    'get_comprehensive_school_analysis': SchoolAnalysisSkill.get_comprehensive_school_analysis,
    'format_school_report': SchoolAnalysisSkill.format_comprehensive_report,
}

SKILL_DESCRIPTIONS = {
    'get_comprehensive_school_analysis': '获取学校综合信息分析（历年分数线、专业组、招生计划、专业详情等）（参数: school_name）',
    'format_school_report': '格式化学校综合报告为可读文本（参数: data）',
}
