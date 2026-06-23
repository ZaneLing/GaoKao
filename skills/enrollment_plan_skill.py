import openpyxl
import os

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class EnrollmentPlanSkill:
    @staticmethod
    def get_plan_by_year(year):
        """按年份获取招生计划"""
        year_map = {
            2021: '招生计划-2021.xlsx',
            2022: '招生计划-2022.xlsx',
            2023: '招生计划-2023-本科.xlsx',
            2024: '招生计划-2024.xlsx',
            2025: '招生计划-2025.xlsx'
        }
        if year not in year_map:
            return []
        
        file_path = os.path.join(DATA_DIR, f'1_最近年份数据/招生计划/{year_map[year]}')
        if not os.path.exists(file_path):
            return []
        
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))[:50]
    
    @staticmethod
    def get_plan_trend(school_name=None, major_name=None, years=None):
        """获取招生计划多年趋势"""
        if years is None:
            years = [2022, 2023, 2024, 2025]
        
        trend_data = {}
        for year in years:
            plan_data = EnrollmentPlanSkill.get_plan_by_year(year)
            if plan_data:
                filtered = []
                for row in plan_data:
                    if row[0] is not None:
                        if school_name and school_name not in str(row[0]):
                            continue
                        if major_name and major_name not in str(row):
                            continue
                        filtered.append(row)
                trend_data[year] = filtered[:10]
        
        return trend_data
    
    @staticmethod
    def analyze_plan_change(school_name, major_name=None):
        """分析招生计划变化趋势"""
        trend_data = EnrollmentPlanSkill.get_plan_trend(school_name=school_name, major_name=major_name)
        
        if not trend_data:
            return {'school': school_name, 'error': '未找到招生计划数据'}
        
        plan_counts = {}
        for year, data in trend_data.items():
            count = 0
            for row in data:
                for val in row:
                    if isinstance(val, int) and val > 0 and val < 1000:
                        count += val
            plan_counts[year] = count
        
        if plan_counts:
            years = sorted(plan_counts.keys())
            first_year = years[0]
            last_year = years[-1]
            
            total_change = plan_counts[last_year] - plan_counts[first_year]
            change_rate = total_change / plan_counts[first_year] * 100 if plan_counts[first_year] > 0 else 0
            
            return {
                'school': school_name,
                'major': major_name if major_name else '全部专业',
                'plan_by_year': plan_counts,
                'total_change': total_change,
                'change_rate': round(change_rate, 1),
                'trend': '扩招' if total_change > 0 else '缩招' if total_change < 0 else '稳定',
                'years_covered': len(years)
            }
        
        return {'school': school_name, 'error': '无法提取计划数据'}
    
    @staticmethod
    def get_school_plan_detail(school_name, year=2024):
        """获取某院校某年度详细招生计划"""
        plan_data = EnrollmentPlanSkill.get_plan_by_year(year)
        
        result = []
        for row in plan_data:
            if row[0] is not None and school_name in str(row[0]):
                result.append(row)
                if len(result) >= 15:
                    break
        
        return result
    
    @staticmethod
    def compare_plan_between_schools(schools, year=2024):
        """对比多所院校的招生计划"""
        plan_data = EnrollmentPlanSkill.get_plan_by_year(year)
        
        comparison = {}
        for school in schools:
            school_plans = []
            for row in plan_data:
                if row[0] is not None and school in str(row[0]):
                    school_plans.append(row)
            comparison[school] = school_plans[:5]
        
        return comparison
    
    @staticmethod
    def analyze_major_plan_distribution(year=2024):
        """分析各专业招生计划分布"""
        plan_data = EnrollmentPlanSkill.get_plan_by_year(year)
        
        major_dist = {}
        for row in plan_data:
            for i, val in enumerate(row):
                if isinstance(val, str) and len(val) > 2 and val not in ['合计', '总计', '小计']:
                    major_name = val
                    if i + 1 < len(row) and isinstance(row[i+1], int):
                        count = row[i+1]
                        major_dist[major_name] = major_dist.get(major_name, 0) + count
        
        sorted_dist = sorted(major_dist.items(), key=lambda x: x[1], reverse=True)
        return sorted_dist[:20]
    
    @staticmethod
    def get_plan_by_major(major_name, year=2024):
        """按专业查询招生计划"""
        plan_data = EnrollmentPlanSkill.get_plan_by_year(year)
        
        result = []
        for row in plan_data:
            if row[0] is not None and major_name in str(row):
                result.append(row)
                if len(result) >= 15:
                    break
        
        return result

SKILLS_MAP = {
    'get_plan_by_year': EnrollmentPlanSkill.get_plan_by_year,
    'get_plan_trend': EnrollmentPlanSkill.get_plan_trend,
    'analyze_plan_change': EnrollmentPlanSkill.analyze_plan_change,
    'get_school_plan_detail': EnrollmentPlanSkill.get_school_plan_detail,
    'compare_plan_between_schools': EnrollmentPlanSkill.compare_plan_between_schools,
    'analyze_major_plan_distribution': EnrollmentPlanSkill.analyze_major_plan_distribution,
    'get_plan_by_major': EnrollmentPlanSkill.get_plan_by_major,
}

SKILL_DESCRIPTIONS = {
    'get_plan_by_year': '按年份获取招生计划',
    'get_plan_trend': '获取招生计划多年趋势',
    'analyze_plan_change': '分析招生计划变化趋势（扩招/缩招）',
    'get_school_plan_detail': '获取某院校某年度详细招生计划',
    'compare_plan_between_schools': '对比多所院校的招生计划',
    'analyze_major_plan_distribution': '分析各专业招生计划分布',
    'get_plan_by_major': '按专业查询招生计划',
}
