import openpyxl
import os

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class MajorGroupSkill:
    _score_cache_2024 = None
    _score_cache_2025 = None
    _plan_cache_2024 = None
    _plan_cache_2025 = None
    
    @staticmethod
    def _load_score_data(year=2024):
        if year == 2025:
            if MajorGroupSkill._score_cache_2025 is None:
                file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/专业分数-2025.xlsx')
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb[wb.sheetnames[0]]
                    MajorGroupSkill._score_cache_2025 = list(ws.iter_rows(values_only=True))
                    print(f"  [缓存] 已加载2025分数数据 {len(MajorGroupSkill._score_cache_2025)} 行")
                except Exception as e:
                    print(f"  [错误] 加载2025分数数据失败: {e}")
                    MajorGroupSkill._score_cache_2025 = []
            return MajorGroupSkill._score_cache_2025
        else:
            if MajorGroupSkill._score_cache_2024 is None:
                file_path = os.path.join(DATA_DIR, '1_最近年份数据/分数线/专业分数-2024-考试院.xlsx')
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb[wb.sheetnames[0]]
                    MajorGroupSkill._score_cache_2024 = list(ws.iter_rows(values_only=True))
                    print(f"  [缓存] 已加载2024分数数据 {len(MajorGroupSkill._score_cache_2024)} 行")
                except Exception as e:
                    print(f"  [错误] 加载2024分数数据失败: {e}")
                    MajorGroupSkill._score_cache_2024 = []
            return MajorGroupSkill._score_cache_2024
    
    @staticmethod
    def _load_plan_data(year=2024):
        if year == 2025:
            if MajorGroupSkill._plan_cache_2025 is None:
                file_path = os.path.join(DATA_DIR, '1_最近年份数据/招生计划/招生计划-2025.xlsx')
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb[wb.sheetnames[0]]
                    MajorGroupSkill._plan_cache_2025 = list(ws.iter_rows(values_only=True))
                    print(f"  [缓存] 已加载2025计划数据 {len(MajorGroupSkill._plan_cache_2025)} 行")
                except Exception as e:
                    print(f"  [错误] 加载2025计划数据失败: {e}")
                    MajorGroupSkill._plan_cache_2025 = []
            return MajorGroupSkill._plan_cache_2025
        else:
            if MajorGroupSkill._plan_cache_2024 is None:
                file_path = os.path.join(DATA_DIR, '1_最近年份数据/招生计划/招生计划-2024.xlsx')
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb[wb.sheetnames[0]]
                    MajorGroupSkill._plan_cache_2024 = list(ws.iter_rows(values_only=True))
                    print(f"  [缓存] 已加载2024计划数据 {len(MajorGroupSkill._plan_cache_2024)} 行")
                except Exception as e:
                    print(f"  [错误] 加载2024计划数据失败: {e}")
                    MajorGroupSkill._plan_cache_2024 = []
            return MajorGroupSkill._plan_cache_2024
    
    @staticmethod
    def get_major_group_scores(major_group=None, school_name=None, major_names=None, subject=None, year=2024):
        data = MajorGroupSkill._load_score_data(year)
        results = []
        
        target_group_num = MajorGroupSkill.get_group_number(major_group) if major_group else None
        
        for row in data[1:]:
            if row[0] is None:
                continue
            if year == 2025:
                row_school = str(row[3]) if row[3] else ''
                row_major = str(row[11]) if row[11] else ''
                row_group = str(row[8]) if row[8] else ''
            else:
                row_school = str(row[1]) if row[1] else ''
                row_major = str(row[6]) if row[6] else ''
                row_group = str(row[13]) if row[13] else ''
            
            if school_name and school_name not in row_school:
                continue
            
            if major_group:
                row_group_num = MajorGroupSkill.get_group_number(row_group)
                if target_group_num and row_group_num and target_group_num == row_group_num:
                    pass
                elif major_group in row_group:
                    pass
                else:
                    continue
            
            if major_names:
                matched = False
                for m in major_names:
                    if m in row_major:
                        matched = True
                        break
                if not matched:
                    continue
            results.append(row)
            if len(results) >= 50:
                break
        return results
    
    @staticmethod
    def get_major_group_plan(major_group=None, school_name=None, year=2024):
        data = MajorGroupSkill._load_plan_data(year)
        results = []
        
        target_group_num = MajorGroupSkill.get_group_number(major_group) if major_group else None
        
        for row in data[1:]:
            if row[0] is None:
                continue
            row_school = str(row[1]) if row[1] else ''
            row_group = str(row[3]) if row[3] else ''
            
            if school_name and school_name not in row_school:
                continue
            
            if major_group:
                row_group_num = MajorGroupSkill.get_group_number(row_group)
                if target_group_num and row_group_num and target_group_num == row_group_num:
                    pass
                elif major_group in row_group:
                    pass
                else:
                    continue
            
            results.append(row)
            if len(results) >= 30:
                break
        return results
    
    @staticmethod
    def extract_group_code(group_name):
        if group_name:
            if '(' in group_name and ')' in group_name:
                inner = group_name[group_name.index('(')+1:group_name.rindex(')')]
                return inner
        return str(group_name)
    
    @staticmethod
    def get_group_number(group_name):
        import re
        match = re.search(r'(\d{2,4})', str(group_name))
        return match.group(1) if match else None
    
    @staticmethod
    def analyze_school_level_metrics(school_name):
        metrics = {
            'school_name': school_name,
            '2024': {
                'min_rank': None,
                'max_rank': None,
                'avg_rank': None,
                'min_score': None,
                'avg_score': None,
                'total_plan': 0,
                'group_count': 0,
                'groups': [],
                'data_count': 0,
                'special_plans': [],
                'school_codes': [],
                'total_enroll': None,
                'group_plans': {}
            },
            '2025': {
                'min_rank': None,
                'max_rank': None,
                'avg_rank': None,
                'min_score': None,
                'avg_score': None,
                'total_plan': 0,
                'group_count': 0,
                'groups': [],
                'data_count': 0,
                'special_plans': [],
                'school_codes': [],
                'total_enroll': None,
                'group_plans': {}
            },
            'comparison': {}
        }
        
        special_plan_keywords = ['中外合作', '中外合办', '国际班', '地方专项', '国家专项', '定向', '公费']
        
        for year in [2024, 2025]:
            try:
                scores = MajorGroupSkill.get_major_group_scores(school_name=school_name, year=year)
                plans = MajorGroupSkill.get_major_group_plan(school_name=school_name, year=year)
            except Exception as e:
                print(f"    警告: 获取{year}年数据失败: {e}")
                continue
            
            if year == 2025:
                ranks = []
                avg_ranks = []
                min_scores = []
                avg_scores = []
                groups = set()
                group_plans = {}
                total_plan = 0
                special_plans = set()
                school_codes = set()
                campus_info = set()
                enroll_counts = []
                
                for row in scores:
                    if len(row) >= 17:
                        if row[15] and isinstance(row[15], (int, float)):
                            ranks.append(row[15])
                        if row[17] and isinstance(row[17], (int, float)):
                            avg_ranks.append(row[17])
                        if row[14] and isinstance(row[14], (int, float)):
                            min_scores.append(row[14])
                        if row[16] and isinstance(row[16], (int, float)):
                            avg_scores.append(row[16])
                        if row[8]:
                            groups.add(str(row[8]))
                        if row[3] and isinstance(row[3], (int, float)):
                            school_codes.add(str(row[3]))
                        if row[20] and isinstance(row[20], (int, float)):
                            enroll_counts.append(row[20])
                        if row[12]:
                            major_note = str(row[12])
                            for kw in special_plan_keywords:
                                if kw in major_note:
                                    special_plans.add(kw)
                
                for row in plans:
                    if len(row) >= 7:
                        if row[6] and isinstance(row[6], (int, float)):
                            total_plan += row[6]
                        if row[3]:
                            group_key = str(row[3])
                            if group_key not in group_plans:
                                group_plans[group_key] = 0
                            if row[12] and isinstance(row[12], (int, float)):
                                group_plans[group_key] += row[12]
                        if row[2]:
                            school_codes.add(str(row[2]))
                        if row[7]:
                            major_str = str(row[7])
                            for kw in special_plan_keywords:
                                if kw in major_str:
                                    special_plans.add(kw)
                
                metrics[year] = {
                    'min_rank': min(ranks) if ranks else None,
                    'max_rank': max(ranks) if ranks else None,
                    'avg_rank': round(sum(avg_ranks)/len(avg_ranks)) if avg_ranks else (round(sum(ranks)/len(ranks)) if ranks else None),
                    'min_score': min(min_scores) if min_scores else None,
                    'avg_score': round(sum(avg_scores)/len(avg_scores)) if avg_scores else (round(sum(min_scores)/len(min_scores)) if min_scores else None),
                    'total_plan': total_plan,
                    'group_count': len(groups),
                    'groups': list(groups)[:10],
                    'data_count': len(scores),
                    'special_plans': list(special_plans),
                    'school_codes': list(school_codes),
                    'total_enroll': sum(enroll_counts) if enroll_counts else None,
                    'group_plans': group_plans
                }
            else:
                ranks = []
                scores_list = []
                groups = set()
                group_plans = {}
                total_plan = 0
                special_plans = set()
                school_codes = set()
                enroll_counts = []
                
                for row in scores:
                    if len(row) >= 14:
                        if row[12] and isinstance(row[12], (int, float)):
                            ranks.append(row[12])
                        elif row[9] and isinstance(row[9], (int, float)):
                            ranks.append(row[9])
                        if row[11] and isinstance(row[11], (int, float)):
                            scores_list.append(row[11])
                        elif row[8] and isinstance(row[8], (int, float)):
                            scores_list.append(row[8])
                        if row[13]:
                            groups.add(str(row[13]))
                        if row[2] and isinstance(row[2], (int, float)):
                            school_codes.add(str(row[2]))
                        if row[14]:
                            note = str(row[14])
                            for kw in special_plan_keywords:
                                if kw in note:
                                    special_plans.add(kw)
                
                for row in plans:
                    if len(row) >= 7:
                        if row[6] and isinstance(row[6], (int, float)):
                            total_plan += row[6]
                        if row[3]:
                            group_key = str(row[3])
                            if group_key not in group_plans:
                                group_plans[group_key] = 0
                            if row[12] and isinstance(row[12], (int, float)):
                                group_plans[group_key] += row[12]
                        if row[2]:
                            school_codes.add(str(row[2]))
                        if row[7]:
                            major_str = str(row[7])
                            for kw in special_plan_keywords:
                                if kw in major_str:
                                    special_plans.add(kw)
                
                metrics[year] = {
                    'min_rank': min(ranks) if ranks else None,
                    'max_rank': max(ranks) if ranks else None,
                    'avg_rank': round(sum(ranks)/len(ranks)) if ranks else None,
                    'min_score': min(scores_list) if scores_list else None,
                    'avg_score': round(sum(scores_list)/len(scores_list)) if scores_list else None,
                    'total_plan': total_plan,
                    'group_count': len(groups),
                    'groups': list(groups)[:10],
                    'data_count': len(scores),
                    'special_plans': list(special_plans),
                    'school_codes': list(school_codes),
                    'group_plans': group_plans
                }
        
        if metrics['2024']['total_plan'] and metrics['2025']['total_plan']:
            plan_diff = metrics['2025']['total_plan'] - metrics['2024']['total_plan']
            plan_ratio = plan_diff / metrics['2024']['total_plan'] * 100
            if plan_ratio > 10:
                metrics['comparison']['enrollment_change'] = f"扩招 {plan_ratio:.1f}%"
            elif plan_ratio < -10:
                metrics['comparison']['enrollment_change'] = f"缩招 {abs(plan_ratio):.1f}%"
            else:
                metrics['comparison']['enrollment_change'] = f"基本持平 ({plan_ratio:.1f}%)"
        
        if metrics['2024']['group_count'] and metrics['2025']['group_count']:
            group_diff = metrics['2025']['group_count'] - metrics['2024']['group_count']
            groups_2024 = set(metrics['2024']['groups'])
            groups_2025 = set(metrics['2025']['groups'])
            new_groups = groups_2025 - groups_2024
            removed_groups = groups_2024 - groups_2025
            
            if group_diff > 0:
                metrics['comparison']['group_change'] = f"新增{group_diff}个专业组"
                if new_groups:
                    metrics['comparison']['new_groups'] = list(new_groups)[:5]
            elif group_diff < 0:
                metrics['comparison']['group_change'] = f"减少{abs(group_diff)}个专业组"
                if removed_groups:
                    metrics['comparison']['removed_groups'] = list(removed_groups)[:5]
            else:
                metrics['comparison']['group_change'] = "专业组数量不变"
        
        if metrics['2024']['school_codes'] and metrics['2025']['school_codes']:
            codes_2024 = set(metrics['2024']['school_codes'])
            codes_2025 = set(metrics['2025']['school_codes'])
            if codes_2024 != codes_2025:
                metrics['comparison']['code_change'] = "招生代码发生变化"
                metrics['comparison']['codes_2024'] = list(codes_2024)
                metrics['comparison']['codes_2025'] = list(codes_2025)
            else:
                metrics['comparison']['code_change'] = "招生代码未变化"
        
        if metrics['2024']['avg_rank'] and metrics['2025']['avg_rank']:
            rank_diff = metrics['2025']['avg_rank'] - metrics['2024']['avg_rank']
            if rank_diff > 8000:
                metrics['comparison']['rank_trend'] = "位次大幅后退（明显大小年）"
            elif rank_diff > 3000:
                metrics['comparison']['rank_trend'] = "位次后退（存在大小年）"
            elif rank_diff > 1000:
                metrics['comparison']['rank_trend'] = "位次略有后退"
            elif rank_diff < -8000:
                metrics['comparison']['rank_trend'] = "位次大幅前进（明显大小年）"
            elif rank_diff < -3000:
                metrics['comparison']['rank_trend'] = "位次前进（存在大小年）"
            elif rank_diff < -1000:
                metrics['comparison']['rank_trend'] = "位次略有前进"
            else:
                metrics['comparison']['rank_trend'] = "位次基本稳定"
        
        special_plans_2024 = set(metrics['2024']['special_plans'])
        special_plans_2025 = set(metrics['2025']['special_plans'])
        all_special_plans = special_plans_2024.union(special_plans_2025)
        if all_special_plans:
            metrics['comparison']['special_plans'] = list(all_special_plans)
            if '中外合作' in all_special_plans:
                metrics['comparison']['has_international'] = True
            if '地方专项' in all_special_plans or '国家专项' in all_special_plans:
                metrics['comparison']['has_special_quota'] = True
            if '定向' in all_special_plans or '公费' in all_special_plans:
                metrics['comparison']['has_定向'] = True
        
        return metrics
    
    @staticmethod
    def analyze_major_level_metrics(school_name, major_group, major_names=None):
        metrics = {
            'school_name': school_name,
            'major_group': major_group,
            '2024': {
                'min_rank': None,
                'max_rank': None,
                'avg_rank': None,
                'min_score': None,
                'max_score': None,
                'avg_score': None,
                'total_plan': 0,
                'majors': [],
                'data_count': 0
            },
            '2025': {
                'min_rank': None,
                'max_rank': None,
                'avg_rank': None,
                'min_score': None,
                'max_score': None,
                'avg_score': None,
                'total_plan': 0,
                'majors': [],
                'data_count': 0
            },
            'comparison': {}
        }
        
        for year in [2024, 2025]:
            year_key = str(year)
            try:
                scores = MajorGroupSkill.get_major_group_scores(
                    school_name=school_name, 
                    major_group=major_group, 
                    major_names=major_names,
                    year=year
                )
                plans = MajorGroupSkill.get_major_group_plan(
                    school_name=school_name, 
                    major_group=major_group,
                    year=year
                )
            except Exception as e:
                print(f"    警告: 获取{year}年数据失败: {e}")
                continue
            
            if year == 2025:
                ranks = []
                avg_ranks = []
                scores_list = []
                max_scores = []
                avg_scores = []
                majors = set()
                total_plan = 0
                
                for row in scores:
                    if len(row) >= 19:
                        if row[15] and isinstance(row[15], (int, float)):
                            ranks.append(row[15])
                        if row[17] and isinstance(row[17], (int, float)):
                            avg_ranks.append(row[17])
                        if row[14] and isinstance(row[14], (int, float)):
                            scores_list.append(row[14])
                        if row[18] and isinstance(row[18], (int, float)):
                            max_scores.append(row[18])
                        if row[16] and isinstance(row[16], (int, float)):
                            avg_scores.append(row[16])
                        if row[11]:
                            majors.add(str(row[11]))
                
                for row in plans:
                    if len(row) >= 7 and row[6] and isinstance(row[6], (int, float)):
                        total_plan += row[6]
                
                metrics[year_key] = {
                    'min_rank': min(ranks) if ranks else None,
                    'max_rank': max(ranks) if ranks else None,
                    'avg_rank': round(sum(avg_ranks)/len(avg_ranks)) if avg_ranks else (round(sum(ranks)/len(ranks)) if ranks else None),
                    'min_score': min(scores_list) if scores_list else None,
                    'max_score': max(max_scores) if max_scores else None,
                    'avg_score': round(sum(avg_scores)/len(avg_scores)) if avg_scores else (round(sum(scores_list)/len(scores_list)) if scores_list else None),
                    'total_plan': total_plan,
                    'majors': list(majors),
                    'data_count': len(scores)
                }
            else:
                ranks = []
                avg_ranks = []
                scores_list = []
                max_scores = []
                avg_scores = []
                majors = set()
                total_plan = 0
                
                for row in scores:
                    if len(row) >= 14:
                        if row[12] and isinstance(row[12], (int, float)):
                            ranks.append(row[12])
                        if row[9] and isinstance(row[9], (int, float)):
                            avg_ranks.append(row[9])
                        if row[11] and isinstance(row[11], (int, float)):
                            scores_list.append(row[11])
                        if row[8] and isinstance(row[8], (int, float)):
                            max_scores.append(row[8])
                        if row[10] and isinstance(row[10], (int, float)):
                            avg_scores.append(row[10])
                        if row[6]:
                            majors.add(str(row[6]))
                
                for row in plans:
                    if len(row) >= 7 and row[6] and isinstance(row[6], (int, float)):
                        total_plan += row[6]
                
                metrics[year_key] = {
                    'min_rank': min(ranks) if ranks else None,
                    'max_rank': max(ranks) if ranks else None,
                    'avg_rank': round(sum(avg_ranks)/len(avg_ranks)) if avg_ranks else (round(sum(ranks)/len(ranks)) if ranks else None),
                    'min_score': min(scores_list) if scores_list else None,
                    'max_score': max(max_scores) if max_scores else None,
                    'avg_score': round(sum(avg_scores)/len(avg_scores)) if avg_scores else (round(sum(scores_list)/len(scores_list)) if scores_list else None),
                    'total_plan': total_plan,
                    'majors': list(majors),
                    'data_count': len(scores)
                }
        
        if metrics['2024']['total_plan'] and metrics['2025']['total_plan']:
            plan_diff = metrics['2025']['total_plan'] - metrics['2024']['total_plan']
            metrics['comparison']['plan_change'] = plan_diff
        
        if metrics['2024']['avg_rank'] and metrics['2025']['avg_rank']:
            rank_diff = metrics['2025']['avg_rank'] - metrics['2024']['avg_rank']
            metrics['comparison']['rank_diff'] = rank_diff
        
        return metrics
    
    @staticmethod
    def analyze_major_group_admission(major_group, school_name, user_score, user_rank=None, major_names=None):
        metrics_2025 = MajorGroupSkill.analyze_major_level_metrics(school_name, major_group, major_names)
        metrics_2024 = MajorGroupSkill.analyze_major_level_metrics(school_name, major_group, major_names)
        
        plan_info = {}
        plans_2025 = MajorGroupSkill.get_major_group_plan(school_name=school_name, major_group=major_group, year=2025)
        if plans_2025:
            all_majors = []
            for row in plans_2025:
                if len(row) >= 14:
                    all_majors.append(str(row[7]) if row[7] else '')
                    plan_info = {
                        '专业组': MajorGroupSkill.extract_group_code(str(row[3])),
                        '科目要求': row[5] if row[5] else 'N/A',
                        '选科要求': row[13] if len(row) > 13 and row[13] else 'N/A',
                        '计划总数': row[6] if row[6] else 'N/A',
                        '包含专业': '; '.join(all_majors),
                        '学费': row[10] if row[10] else 'N/A',
                        '学制': row[11] if row[11] else 'N/A'
                    }
        
        if not plans_2025:
            plans_2024 = MajorGroupSkill.get_major_group_plan(school_name=school_name, major_group=major_group, year=2024)
            if plans_2024:
                all_majors = []
                for row in plans_2024:
                    if len(row) >= 13:
                        all_majors.append(str(row[7]) if row[7] else '')
                        plan_info = {
                            '专业组': MajorGroupSkill.extract_group_code(str(row[3])),
                            '科目要求': row[5] if row[5] else 'N/A',
                            '计划总数': row[6] if row[6] else 'N/A',
                            '包含专业': '; '.join(all_majors),
                            '学费': row[10] if row[10] else 'N/A',
                            '学制': row[11] if row[11] else 'N/A'
                        }
        
        group_ranks = []
        group_scores = []
        
        if metrics_2025['2025']['avg_rank']:
            group_ranks.append(metrics_2025['2025']['avg_rank'])
            group_ranks.append(metrics_2025['2025']['min_rank'] or metrics_2025['2025']['avg_rank'])
            group_ranks.append(metrics_2025['2025']['max_rank'] or metrics_2025['2025']['avg_rank'])
        
        if metrics_2024['2024']['avg_rank']:
            group_ranks.append(metrics_2024['2024']['avg_rank'])
            group_ranks.append(metrics_2024['2024']['min_rank'] or metrics_2024['2024']['avg_rank'])
            group_ranks.append(metrics_2024['2024']['max_rank'] or metrics_2024['2024']['avg_rank'])
        
        if metrics_2025['2025']['avg_score']:
            group_scores.append(metrics_2025['2025']['avg_score'])
            group_scores.append(metrics_2025['2025']['min_score'] or metrics_2025['2025']['avg_score'])
            group_scores.append(metrics_2025['2025']['max_score'] or metrics_2025['2025']['avg_score'])
        
        if metrics_2024['2024']['avg_score']:
            group_scores.append(metrics_2024['2024']['avg_score'])
            group_scores.append(metrics_2024['2024']['min_score'] or metrics_2024['2024']['avg_score'])
            group_scores.append(metrics_2024['2024']['max_score'] or metrics_2024['2024']['avg_score'])
        
        probability = '未知'
        suggestion = '无法评估'
        level = '不推荐'
        
        if group_ranks and user_rank:
            avg_rank = sum(group_ranks) / len(group_ranks)
            min_rank = min(group_ranks)
            max_rank = max(group_ranks)
            
            rank_diff = user_rank - avg_rank
            
            if user_rank <= min_rank:
                probability = '很高 (90%以上)'
                suggestion = '排名远超最低位次，可以作为保的选择'
                level = '保'
            elif rank_diff <= -5000:
                probability = '较高 (70%-90%)'
                suggestion = '排名优于平均位次，可以作为稳的选择'
                level = '稳'
            elif rank_diff <= 0:
                probability = '一般 (40%-70%)'
                suggestion = '排名接近平均位次，需要谨慎考虑'
                level = '冲'
            elif rank_diff <= 5000:
                probability = '较低 (10%-40%)'
                suggestion = '排名略低于平均位次，冲刺选择'
                level = '冲'
            else:
                probability = '很低 (10%以下)'
                suggestion = '排名差距较大，不建议报考'
                level = '不推荐'
            
            avg_score = sum(group_scores) / len(group_scores) if group_scores else 0
            min_score = min(group_scores) if group_scores else 0
            max_score = max(group_scores) if group_scores else 0
            
            return {
                'school': school_name,
                'major_group': MajorGroupSkill.extract_group_code(major_group),
                'user_score': user_score,
                'user_rank': user_rank,
                'group_avg_score': round(avg_score, 1),
                'group_min_score': min_score,
                'group_max_score': max_score,
                'group_avg_rank': round(avg_rank, 0),
                'group_min_rank': int(min_rank),
                'group_max_rank': int(max_rank),
                'rank_diff': round(rank_diff, 0),
                'probability': probability,
                'suggestion': suggestion,
                'level': level,
                'plan_info': plan_info,
                'school_metrics': MajorGroupSkill.analyze_school_level_metrics(school_name),
                'major_metrics': metrics_2025
            }
        
        elif group_scores:
            avg_score = sum(group_scores) / len(group_scores)
            min_score = min(group_scores)
            max_score = max(group_scores)
            
            score_diff = user_score - avg_score
            
            if user_score >= max_score:
                probability = '很高 (90%以上)'
                suggestion = '可以作为稳或保的选择'
                level = '保'
            elif score_diff >= -5:
                probability = '较高 (70%-90%)'
                suggestion = '可以作为稳的选择'
                level = '稳'
            elif score_diff >= -15:
                probability = '一般 (40%-70%)'
                suggestion = '需要谨慎考虑，可作为冲的选择'
                level = '冲'
            elif score_diff >= -25:
                probability = '较低 (10%-40%)'
                suggestion = '冲刺选择，风险较大'
                level = '冲'
            else:
                probability = '很低 (10%以下)'
                suggestion = '不建议报考'
                level = '不推荐'
            
            return {
                'school': school_name,
                'major_group': MajorGroupSkill.extract_group_code(major_group),
                'user_score': user_score,
                'user_rank': user_rank,
                'group_avg_score': round(avg_score, 1),
                'group_min_score': min_score,
                'group_max_score': max_score,
                'score_diff': round(score_diff, 1),
                'probability': probability,
                'suggestion': suggestion,
                'level': level,
                'plan_info': plan_info,
                'school_metrics': MajorGroupSkill.analyze_school_level_metrics(school_name),
                'major_metrics': metrics_2025
            }
        
        return {
            'school': school_name,
            'major_group': MajorGroupSkill.extract_group_code(major_group),
            'error': '无法提取分数数据',
            'plan_info': plan_info,
            'school_metrics': MajorGroupSkill.analyze_school_level_metrics(school_name),
            'major_metrics': metrics_2025
        }
    
    @staticmethod
    def search_major_groups_by_major(major_name, year=2024):
        data = MajorGroupSkill._load_plan_data(year)
        groups = {}
        for row in data[1:]:
            if row[0] is None or len(row) < 8:
                continue
            school = str(row[1]) if row[1] else ''
            group_full = str(row[3]) if row[3] else ''
            group_code = str(row[2]) if row[2] else ''
            major = str(row[7]) if row[7] else ''
            
            group_number = None
            if '[' in group_code and ']' in group_code:
                group_number = group_code[group_code.index('[')+1:group_code.index(']')]
            if not group_number:
                group_number = MajorGroupSkill.get_group_number(group_full)
            if not group_number:
                group_number = MajorGroupSkill.get_group_number(group_code)
            
            display_group = group_full if group_full else group_code
            
            if major_name in major:
                group_key = f"{school}_{group_number}" if group_number else f"{school}_{display_group}"
                if group_key not in groups:
                    groups[group_key] = {
                        'school': school,
                        'group': display_group,
                        'group_number': group_number,
                        'majors': [],
                        'subject': row[5] if row[5] else 'N/A',
                        'total_plan': row[6] if row[6] else 'N/A'
                    }
                groups[group_key]['majors'].append(major)
        
        return [(g['school'], g) for g in groups.values()]
    
    @staticmethod
    def get_major_groups_by_user_preference(user_score=580, user_rank=None, province='安徽', subject=None, score=None, rank=None, target_majors=None):
        """根据用户指定的专业获取专业组推荐（严格匹配）"""
        # 处理参数别名
        if score and not user_score:
            user_score = score
        if rank and not user_rank:
            user_rank = rank
        
        # 必须指定专业
        if not target_majors:
            print("❌ 错误：必须指定意向专业！")
            return []
        
        # 解析用户指定的专业关键词
        if isinstance(target_majors, list):
            major_keywords = target_majors
        else:
            major_keywords = [m.strip() for m in str(target_majors).split(',')]
        
        print(f"  正在搜索专业组，关键词: {major_keywords}")
        if user_rank:
            print(f"  用户排名: {user_rank}")
        print(f"  ⚠️  所有推荐将严格包含你选择的专业")
        
        all_groups = []
        for i, keyword in enumerate(major_keywords):
            groups = MajorGroupSkill.search_major_groups_by_major(keyword, year=2025)
            all_groups.extend(groups)
            print(f"    关键词 '{keyword}' 找到 {len(groups)} 个专业组")
        
        print(f"  共找到 {len(all_groups)} 个相关专业组（含重复）")
        
        # 去重
        unique_groups = {}
        for school, info in all_groups:
            group_key = f"{school}_{info['group_number']}" if info.get('group_number') else f"{school}_{info['group']}"
            if group_key not in unique_groups:
                unique_groups[group_key] = info
        
        print(f"  去重后 {len(unique_groups)} 个专业组")
        
        # 分析每个专业组的录取可能性
        recommendations = []
        for j, (group_key, info) in enumerate(list(unique_groups.items())[:300]):  # 扩大候选范围到300
            group_param = info['group_number'] if info.get('group_number') else info['group']
            analysis = MajorGroupSkill.analyze_major_group_admission(
                group_param, info['school'], user_score, user_rank, info['majors']
            )
            if 'error' not in analysis:
                # 验证专业组是否包含用户指定的专业
                majors_str = str(info['majors'])
                contains_target = False
                for keyword in major_keywords:
                    if keyword in majors_str:
                        contains_target = True
                        break
                
                if contains_target:
                    recommendations.append(analysis)
                    if 'group_avg_rank' in analysis:
                        print(f"    分析完成: {info['school']} - {analysis['level']} - 平均位次: {analysis['group_avg_rank']}")
                    else:
                        print(f"    分析完成: {info['school']} - {analysis['level']} - 平均分: {analysis['group_avg_score']}")
        
        # 按冲稳保排序
        level_order = {'保': 0, '稳': 1, '冲': 2, '不推荐': 3}
        recommendations.sort(key=lambda x: level_order.get(x.get('level', '不推荐'), 3))
        
        # 确保至少15个，最多45个推荐，按冲稳保比例分配
        min_recommendations = 15
        max_recommendations = 45
        
        # 统计各等级数量
        chong = [r for r in recommendations if r.get('level') == '冲']
        wen = [r for r in recommendations if r.get('level') == '稳']
        bao = [r for r in recommendations if r.get('level') == '保']
        
        print(f"  原始分布: 冲:{len(chong)} 稳:{len(wen)} 保:{len(bao)}")
        
        # 按比例选取：冲:稳:保 = 1:2:2 (共15个) 或 3:6:6 (共15个) 或 9:18:18 (共45个)
        # 目标比例：冲:稳:保 = 1:2:2
        target_chong = min(len(chong), max(5, int(max_recommendations * 0.2)))
        target_wen = min(len(wen), max(10, int(max_recommendations * 0.4)))
        target_bao = min(len(bao), max(10, int(max_recommendations * 0.4)))
        
        # 调整数量确保总数在15-45之间
        total_target = target_chong + target_wen + target_bao
        if total_target < min_recommendations:
            # 如果总数不足，从剩余中补充
            remaining = min_recommendations - total_target
            for r in recommendations:
                if r not in chong[:target_chong] + wen[:target_wen] + bao[:target_bao]:
                    if r.get('level') == '冲' and len(chong) > target_chong:
                        target_chong += 1
                        remaining -= 1
                    elif r.get('level') == '稳' and len(wen) > target_wen:
                        target_wen += 1
                        remaining -= 1
                    elif r.get('level') == '保' and len(bao) > target_bao:
                        target_bao += 1
                        remaining -= 1
                    if remaining <= 0:
                        break
        
        final_recommendations = chong[:target_chong] + wen[:target_wen] + bao[:target_bao]
        
        # 如果还是不足，从所有推荐中补充
        if len(final_recommendations) < min_recommendations:
            for r in recommendations:
                if r not in final_recommendations:
                    final_recommendations.append(r)
                    if len(final_recommendations) >= min_recommendations:
                        break
        
        # 限制最多45个
        final_recommendations = final_recommendations[:max_recommendations]
        
        print(f"  最终推荐: {len(final_recommendations)} 个志愿 (冲:{len([r for r in final_recommendations if r.get('level') == '冲'])} 稳:{len([r for r in final_recommendations if r.get('level') == '稳'])} 保:{len([r for r in final_recommendations if r.get('level') == '保'])})")
        
        return final_recommendations
    
    @staticmethod
    def get_medical_major_groups(user_score=580, user_rank=None, province='安徽', subject=None, score=None, rank=None, target_majors=None):
        # 处理参数别名
        if score and not user_score:
            user_score = score
        if rank and not user_rank:
            user_rank = rank
        
        # 使用用户指定的专业关键词，如果没有则使用默认医学关键词
        if target_majors:
            medical_keywords = target_majors if isinstance(target_majors, list) else [m.strip() for m in str(target_majors).split(',')]
        else:
            medical_keywords = ['临床医学', '口腔医学', '中医学', '药学', '护理', '医学', '生物医学', '针灸推拿', '中西医', '预防医学', '麻醉学', '影像学']
        
        print(f"  正在搜索医学专业组，关键词: {medical_keywords}")
        if user_rank:
            print(f"  用户排名: {user_rank}")
        
        all_groups = []
        for i, keyword in enumerate(medical_keywords):
            groups = MajorGroupSkill.search_major_groups_by_major(keyword, year=2025)
            all_groups.extend(groups)
            print(f"    关键词 '{keyword}' 找到 {len(groups)} 个专业组")
        
        print(f"  共找到 {len(all_groups)} 个医学相关专业组（含重复）")
        
        unique_groups = {}
        for school, info in all_groups:
            group_key = f"{school}_{info['group_number']}" if info.get('group_number') else f"{school}_{info['group']}"
            if group_key not in unique_groups:
                unique_groups[group_key] = info
        
        print(f"  去重后 {len(unique_groups)} 个专业组")
        
        recommendations = []
        for j, (group_key, info) in enumerate(list(unique_groups.items())[:200]):
            group_param = info['group_number'] if info.get('group_number') else info['group']
            analysis = MajorGroupSkill.analyze_major_group_admission(
                group_param, info['school'], user_score, user_rank, info['majors']
            )
            if 'error' not in analysis:
                recommendations.append(analysis)
                if 'group_avg_rank' in analysis:
                    print(f"    分析完成: {info['school']} - {analysis['level']} - 平均位次: {analysis['group_avg_rank']}")
                else:
                    print(f"    分析完成: {info['school']} - {analysis['level']} - 平均分: {analysis['group_avg_score']}")
        
        level_order = {'保': 0, '稳': 1, '冲': 2, '不推荐': 3}
        recommendations.sort(key=lambda x: level_order.get(x.get('level', '不推荐'), 3))
        
        # 确保至少15-20个推荐，按冲稳保比例分配
        min_recommendations = 15
        max_recommendations = 20
        
        # 统计各等级数量
        chong = [r for r in recommendations if r.get('level') == '冲']
        wen = [r for r in recommendations if r.get('level') == '稳']
        bao = [r for r in recommendations if r.get('level') == '保']
        
        # 如果总数不足，按比例补充
        if len(recommendations) < min_recommendations:
            # 目标比例：冲:稳:保 = 3:7:10 (共20个)
            target_chong = max(3, min(len(chong), 5))
            target_wen = max(7, min(len(wen), 10))
            target_bao = max(10, min(len(bao), 15))
            
            # 调整数量
            final_recommendations = chong[:target_chong] + wen[:target_wen] + bao[:target_bao]
            
            # 如果还是不足，从剩余中补充
            remaining = min_recommendations - len(final_recommendations)
            if remaining > 0:
                for r in recommendations:
                    if r not in final_recommendations:
                        final_recommendations.append(r)
                        remaining -= 1
                        if remaining <= 0:
                            break
            
            recommendations = final_recommendations[:max_recommendations]
        
        print(f"  生成 {len(recommendations)} 个推荐结果 (冲:{len([r for r in recommendations if r.get('level') == '冲'])} 稳:{len([r for r in recommendations if r.get('level') == '稳'])} 保:{len([r for r in recommendations if r.get('level') == '保'])})")
        
        return recommendations
    
    @staticmethod
    def compare_major_groups(major_groups, user_score, user_rank=None):
        comparisons = []
        
        for school, group in major_groups:
            analysis = MajorGroupSkill.analyze_major_group_admission(group, school, user_score, user_rank)
            comparisons.append(analysis)
        
        level_order = {'保': 0, '稳': 1, '冲': 2, '不推荐': 3}
        comparisons.sort(key=lambda x: level_order.get(x.get('level', '不推荐'), 3))
        
        return comparisons

SKILLS_MAP = {
    'get_major_group_scores': MajorGroupSkill.get_major_group_scores,
    'get_major_group_plan': MajorGroupSkill.get_major_group_plan,
    'analyze_major_group_admission': MajorGroupSkill.analyze_major_group_admission,
    'search_major_groups_by_major': MajorGroupSkill.search_major_groups_by_major,
    'get_medical_major_groups': MajorGroupSkill.get_medical_major_groups,
    'get_major_groups_by_user_preference': MajorGroupSkill.get_major_groups_by_user_preference,
    'compare_major_groups': MajorGroupSkill.compare_major_groups,
    'analyze_school_level_metrics': MajorGroupSkill.analyze_school_level_metrics,
    'analyze_major_level_metrics': MajorGroupSkill.analyze_major_level_metrics,
}

SKILL_DESCRIPTIONS = {
    'get_major_group_scores': '获取专业组分数线(参数: school_name, major_group, major_names, year, subject)',
    'get_major_group_plan': '获取专业组招生计划(参数: school_name, major_group, year)',
    'analyze_major_group_admission': '分析专业组录取可能性(参数: major_group, school_name, user_score, user_rank, major_names)',
    'search_major_groups_by_major': '根据专业名称搜索专业组(参数: major_name, year)',
    'get_medical_major_groups': '获取医学相关专业组推荐(参数: user_score, user_rank, province, subject, score, rank, target_majors)',
    'get_major_groups_by_user_preference': '根据用户指定专业获取专业组推荐（严格匹配，15-45个）(参数: user_score, user_rank, province, subject, score, rank, target_majors)',
    'compare_major_groups': '对比多个专业组(参数: groups_info)',
    'analyze_school_level_metrics': '分析院校层面指标（位次、计划、扩招等）(参数: school_name, major_group)',
    'analyze_major_level_metrics': '分析专业层面指标（位次、计划、冷热程度等）(参数: school_name, major_group, major_names)',
}
