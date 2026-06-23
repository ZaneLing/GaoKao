import os
import re
from functools import lru_cache

import openpyxl


DATA_DIR = os.environ.get("GAOKAO_DATA_DIR", "/Users/lingziyang/Desktop/Gaokao/安徽")
DEFAULT_PROVINCE = "安徽"


def data_path(relative_path):
    return os.path.join(DATA_DIR, relative_path)


def clean_text(value, default=""):
    if value is None:
        return default
    text = str(value).strip()
    if text in {"", "-", "—", "None", "none", "nan"}:
        return default
    return text


def to_int(value, default=None):
    if value is None:
        return default
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else int(round(value))
    text = clean_text(value)
    if not text:
        return default
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return default
    number = float(match.group(0))
    return int(number) if number.is_integer() else int(round(number))


def to_float(value, default=None):
    if value is None:
        return default
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = clean_text(value)
    if not text:
        return default
    text = text.replace(",", "").replace("%", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else default


def normalize_year(value):
    return to_int(value)


def normalize_province(value, default=DEFAULT_PROVINCE):
    text = clean_text(value)
    if not text:
        return default
    text = re.sub(r"^[A-Z]", "", text)
    return text


def province_matches(value, province=DEFAULT_PROVINCE):
    text = normalize_province(value)
    return not province or province in text


def normalize_subject(value):
    text = clean_text(value)
    if not text:
        return ""
    if "物理" in text or text in {"物", "理科"} or text.startswith("理科"):
        return "物理"
    if "历史" in text or text in {"史", "文科"} or text.startswith("文科"):
        return "历史"
    return text


def normalize_subject_requirement(subject, requirement):
    req = clean_text(requirement)
    if not req:
        return ""
    req = req.replace("物理", "物").replace("历史", "史")
    req = req.replace("化学", "化").replace("生物", "生").replace("思想政治", "政").replace("政治", "政")
    req = req.replace("、", "+").replace("/", "+")
    if "+" in req:
        return req

    primary = normalize_subject(subject)
    prefix = "物" if primary == "物理" else "史" if primary == "历史" else ""
    if prefix:
        return f"{prefix}+{req}"
    return req


def subject_matches(value, subject):
    if not subject:
        return True
    return normalize_subject(value) == normalize_subject(subject)


def get_any(row, names, default=None):
    for name in names:
        if name in row and row[name] not in (None, ""):
            return row[name]
    return default


def _dedupe_headers(headers):
    result = []
    counts = {}
    for idx, header in enumerate(headers):
        name = clean_text(header, f"列{idx}")
        count = counts.get(name, 0)
        counts[name] = count + 1
        if count:
            name = f"{name}_{count + 1}"
        result.append(name)
    return tuple(result)


@lru_cache(maxsize=128)
def sheet_names(relative_path):
    wb = openpyxl.load_workbook(data_path(relative_path), read_only=True, data_only=True)
    return tuple(wb.sheetnames)


def select_sheet(relative_path, sheet_name=None, province=None, year=None):
    names = sheet_names(relative_path)
    if sheet_name and sheet_name in names:
        return sheet_name
    if province and province in names:
        return province
    if year and str(year) in names:
        return str(year)
    return names[0]


@lru_cache(maxsize=256)
def read_sheet_rows(relative_path, sheet_name=None):
    selected_sheet = select_sheet(relative_path, sheet_name=sheet_name)
    wb = openpyxl.load_workbook(data_path(relative_path), read_only=True, data_only=True)
    ws = wb[selected_sheet]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = _dedupe_headers(next(iterator))
    except StopIteration:
        return tuple()

    rows = []
    for raw in iterator:
        if not raw or all(value is None for value in raw):
            continue
        rows.append({headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))})
    return tuple(rows)


def read_rows(relative_path, sheet_name=None, province_sheet=None, year_sheet=None, all_sheets=False):
    if all_sheets:
        rows = []
        for name in sheet_names(relative_path):
            rows.extend(read_sheet_rows(relative_path, name))
        return rows
    selected = select_sheet(relative_path, sheet_name=sheet_name, province=province_sheet, year=year_sheet)
    return list(read_sheet_rows(relative_path, selected))


def extract_school_code(admission_code):
    text = clean_text(admission_code)
    if not text:
        return ""
    match = re.match(r"(\d+)", text)
    return match.group(1) if match else text


def extract_group_code(*values):
    for value in values:
        text = clean_text(value)
        if not text:
            continue
        bracket = re.search(r"\[(\d{2,4})\]", text)
        if bracket:
            return bracket.group(1).zfill(3)
        major_group = re.search(r"(\d{2,4})\s*专业组", text)
        if major_group:
            return major_group.group(1).zfill(3)
        if re.fullmatch(r"\d{1,4}", text):
            return text.zfill(3) if len(text) <= 3 else text
    return ""


def build_major_group_key(year, batch, subject, admission_code, group_code, subject_requirement):
    parts = [
        clean_text(year),
        clean_text(batch),
        normalize_subject(subject),
        clean_text(admission_code),
        clean_text(group_code),
        clean_text(subject_requirement),
    ]
    return "|".join(parts)


def normalize_batch_line(row, province=DEFAULT_PROVINCE):
    province_value = normalize_province(get_any(row, ["省市", "省份"], province))
    return {
        "省份": province_value,
        "年份": normalize_year(row.get("年份")),
        "批次": clean_text(get_any(row, ["批次/段", "批次"])),
        "科类": normalize_subject(get_any(row, ["科目", "科类"])),
        "批次类型": clean_text(row.get("批次类型")),
        "控制分数线": to_int(row.get("控制分数线")),
        "压分线": to_int(row.get("压分线")),
        "压线分区间": clean_text(row.get("压线分区间")),
        "备注": clean_text(get_any(row, ["remark", "备注"])),
    }


def normalize_score_segment(row, province=DEFAULT_PROVINCE):
    return {
        "省份": normalize_province(row.get("省份"), province),
        "年份": normalize_year(row.get("年份")),
        "科类": normalize_subject(row.get("科目")),
        "层次": clean_text(row.get("层次")),
        "最高位次": to_int(row.get("最高位次")),
        "最低位次": to_int(row.get("最低位次")),
        "最大分数": to_int(row.get("最大分数")),
        "最小分数": to_int(row.get("最小分数")),
        "分数区间": clean_text(row.get("分数区间")),
        "同分人数": to_int(row.get("同分人数")),
    }


def normalize_school_score(row, province=DEFAULT_PROVINCE):
    return {
        "省份": province,
        "年份": normalize_year(row.get("年份")),
        "学校": clean_text(get_any(row, ["学校", "院校名称"])),
        "招生代码": clean_text(get_any(row, ["招生代码", "院校代码"])),
        "学校方向": clean_text(get_any(row, ["学校方向", "方向名称"])),
        "批次": clean_text(get_any(row, ["批次", "录取批次"])),
        "科类": normalize_subject(get_any(row, ["科目", "科类"])),
        "录取人数": to_int(row.get("录取人数")),
        "最高分": to_int(row.get("最高分")),
        "平均分": to_float(row.get("平均分")),
        "最低分": to_int(row.get("最低分")),
        "最低分位次": to_int(get_any(row, ["最低分位次", "最低位次"])),
        "最低分线差": to_int(row.get("最低分线差")),
    }


def normalize_major_score(row, source_year=None, province=DEFAULT_PROVINCE):
    year = normalize_year(get_any(row, ["年份"], source_year))
    school = clean_text(get_any(row, ["院校名称", "学校"]))
    school_code = clean_text(get_any(row, ["院校代码", "招生代码"]))
    admission_code = clean_text(get_any(row, ["院校专业组代码", "招生代码", "院校代码"], school_code))
    group_name = clean_text(get_any(row, ["专业组", "学校方向", "院校专业组代码"]))
    group_code = extract_group_code(get_any(row, ["专业组代码"]), group_name, admission_code)
    subject = normalize_subject(get_any(row, ["科类", "科目"]))
    subject_requirement = normalize_subject_requirement(subject, get_any(row, ["选科要求", "再选"]))
    batch = clean_text(get_any(row, ["批次", "录取批次"]))

    normalized = {
        "省份": normalize_province(get_any(row, ["生源地", "省份"], province), province),
        "年份": year,
        "院校名称": school,
        "学校": school,
        "院校代码": school_code,
        "招生代码": admission_code,
        "院校专业组代码": clean_text(row.get("院校专业组代码")),
        "专业组代码": group_code,
        "专业组名称": group_name or group_code,
        "批次": batch,
        "科类": subject,
        "批次类别": clean_text(row.get("批次类别")),
        "专业代码": clean_text(row.get("专业代码")),
        "专业全称": clean_text(row.get("专业全称")),
        "专业名称": clean_text(get_any(row, ["专业名称", "专业全称"])),
        "专业备注": clean_text(get_any(row, ["专业备注", "备注"])),
        "选科要求": subject_requirement,
        "最高分": to_int(row.get("最高分")),
        "最高位次": to_int(row.get("最高位次")),
        "平均分": to_float(get_any(row, ["平均分", "均分"])),
        "平均位次": to_int(row.get("平均位次")),
        "最低分": to_int(row.get("最低分")),
        "最低位次": to_int(get_any(row, ["最低位次", "低分位次"])),
        "录取人数": to_int(row.get("录取人数")),
        "备注": clean_text(row.get("备注")),
    }
    normalized["专业组key"] = build_major_group_key(
        year, batch, subject, school_code or admission_code, group_code, subject_requirement
    )
    return normalized


def normalize_plan(row, source_year=None, province=DEFAULT_PROVINCE):
    year = normalize_year(get_any(row, ["年份"], source_year))
    school = clean_text(get_any(row, ["学校", "院校名称"]))
    admission_code = clean_text(get_any(row, ["招生代码", "院码", "院校代码"]))
    school_code = extract_school_code(admission_code)
    group_name = clean_text(get_any(row, ["学校方向", "方向名称"]))
    group_code = extract_group_code(get_any(row, ["专业组代码"]), admission_code, group_name)
    subject = normalize_subject(get_any(row, ["科目", "科类", "文科"]))
    subject_requirement = normalize_subject_requirement(subject, get_any(row, ["选科要求", "选科"]))
    batch = clean_text(get_any(row, ["批次", "录取批次"]))
    province_value = normalize_province(get_any(row, ["省份", "招生地区", "生源地"], province), province)
    plan_count = to_int(get_any(row, ["计划人数", "计划数"]))
    total_plan = to_int(row.get("计划总数"))

    normalized = {
        "省份": province_value,
        "年份": year,
        "学校": school,
        "院校名称": school,
        "院校代码": school_code,
        "招生代码": admission_code,
        "专业组代码": group_code,
        "专业组名称": group_name or (f"{school}{group_code}专业组" if group_code else ""),
        "学校方向": group_name,
        "科类": subject,
        "选科要求": subject_requirement,
        "计划总数": total_plan,
        "专业": clean_text(get_any(row, ["专业", "专业名称"])),
        "专业名称": clean_text(get_any(row, ["专业", "专业名称"])),
        "专业代码": clean_text(get_any(row, ["专业代码", "专码"])),
        "批次": batch,
        "学费": clean_text(row.get("学费")),
        "学制": clean_text(row.get("学制")),
        "计划人数": plan_count,
        "语种": clean_text(row.get("语种")),
        "备注": clean_text(row.get("备注")),
        "enroll_unit_id": clean_text(row.get("enroll_unit_id")),
        "enroll_major_id": clean_text(row.get("enroll_major_id")),
    }
    normalized["专业组key"] = build_major_group_key(
        year, batch, subject, school_code or admission_code, group_code, subject_requirement
    )
    return normalized


def filter_rows(rows, year=None, province=DEFAULT_PROVINCE, subject=None, batch=None):
    result = []
    for row in rows:
        if province and row.get("省份") and not province_matches(row.get("省份"), province):
            continue
        if year and row.get("年份") != int(year):
            continue
        if subject and not subject_matches(row.get("科类"), subject):
            continue
        if batch and batch not in clean_text(row.get("批次")):
            continue
        result.append(row)
    return result
