import openpyxl
import xlrd
import os
from .data_utils import clean_text, read_rows

DATA_DIR = '/Users/lingziyang/Desktop/Gaokao/安徽'

class SchoolInfoSkill:
    @staticmethod
    def get_school_info(school_name=None):
        """获取院校基础信息"""
        file_path = os.path.join(DATA_DIR, '3_院校信息/院校基础信息.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        data = []
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None:
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(row)
                if len(data) >= 10:
                    break
        
        return data
    
    @staticmethod
    def get_school_detailed_info(school_name):
        """获取院校详细信息（含保研率、硕士点、博士点等）"""
        file_path = os.path.join(DATA_DIR, '3_院校信息/院校基础信息.xlsx')
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb[wb.sheetnames[0]]
        
        headers = next(ws.iter_rows(values_only=True))
        
        for row in ws.iter_rows(values_only=True):
            if row[0] is not None and school_name in str(row[0]):
                info = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        info[str(header)] = row[i]
                
                return {
                    '学校名称': info.get('学校名称'),
                    '新院校名称': info.get('新院校名称'),
                    '排名': info.get('排名'),
                    '所在省': info.get('所在省'),
                    '城市': info.get('城市'),
                    '类型': info.get('类型'),
                    '隶属单位': info.get('隶属单位'),
                    '是否985': info.get('是否985'),
                    '是否211': info.get('是否211'),
                    '一流大学': info.get('一流大学'),
                    '是否艺术': info.get('是否艺术'),
                    '国重/省重': info.get('国重/省重'),
                    '公私性质': info.get('公私性质'),
                    '本科/专科': info.get('本科/专科'),
                    '保研率': info.get('保研率'),
                    '国家特色专业': info.get('国家特色专业'),
                    '省特色专业': info.get('省特色专业'),
                    '是否国重点': info.get('是否国重点'),
                    '世界一流': info.get('世界一流'),
                    '是否双一流': info.get('是否双一流'),
                    '硕士点（个）': info.get('硕士点（个）'),
                    '博士点（个）': info.get('博士点（个）'),
                    '成立时间': info.get('成立时间'),
                    '女生比例': info.get('女生比例'),
                    '男生比例': info.get('男生比例'),
                    '招办电话': info.get('招办电话'),
                    '电子邮箱': info.get('电子邮箱'),
                    '通讯地址': info.get('通讯地址'),
                    '官网': info.get('官网'),
                    '评估结果': info.get('评估结果'),
                    '大学简介': str(info.get('大学简介'))[:300] if info.get('大学简介') else None
                }
        
        return None
    
    @staticmethod
    def get_school_campus_info(school_name):
        """获取院校校区信息"""
        detailed_info = SchoolInfoSkill.get_school_detailed_info(school_name)
        
        if not detailed_info:
            return {'school': school_name, 'error': '未找到院校信息'}
        
        address = detailed_info.get('通讯地址', '')
        city = detailed_info.get('城市', '')
        
        campus_info = {
            'school': school_name,
            'city': city,
            'address': address,
            'has_multiple_campuses': False,
            'campus_count': 1
        }
        
        if address and ('校区' in address or '分校' in address):
            campus_info['has_multiple_campuses'] = True
        
        return campus_info
    
    @staticmethod
    def analyze_school_medical_strength(school_name):
        """分析院校医学学科实力"""
        discipline_data = SchoolInfoSkill.get_discipline_evaluation(school_name=school_name)
        
        medical_disciplines = ['临床医学', '基础医学', '口腔医学', '公共卫生与预防医学', 
                              '中医学', '中西医结合', '药学', '中药学', '护理学']
        
        found_disciplines = []
        a_count = 0
        b_count = 0
        c_count = 0
        
        for row in discipline_data:
            discipline_name = clean_text(row.get('专业类'))
            level = clean_text(row.get('评估结果'))
            if any(md in discipline_name for md in medical_disciplines):
                found_disciplines.append(f"{discipline_name}: {level}")
                if 'A' in level:
                    a_count += 1
                elif 'B' in level:
                    b_count += 1
                elif 'C' in level:
                    c_count += 1
        
        strength_level = '强' if a_count >= 2 else '较强' if b_count >= 2 else '一般'
        
        return {
            'school': school_name,
            'medical_disciplines': found_disciplines,
            'A_level_count': a_count,
            'B_level_count': b_count,
            'C_level_count': c_count,
            'total_medical_disciplines': len(found_disciplines),
            'strength_level': strength_level
        }
    
    @staticmethod
    def get_school_ranking(school_name=None, ranking_type='综合'):
        """获取院校排名，支持多种排名类型综合对比"""
        file_map = {
            '软科': '2022排名_软科.xlsx',
            '校友会': '2022排名_校友会.xlsx',
            'QS': '2022排名_QS.xlsx',
            'U.S.News': '2022排名_U.S.News.xlsx',
            '泰晤士': '2022排名_泰晤士.xlsx'
        }
        
        if ranking_type == '综合':
            results = {}
            for r_type, file in file_map.items():
                file_path = os.path.join(DATA_DIR, f'5_参考数据/{file}')
                try:
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb[wb.sheetnames[0]]
                    for row in ws.iter_rows(values_only=True):
                        if row[0] is not None:
                            if school_name and school_name not in str(row[0]):
                                continue
                            results[r_type] = {
                                'school': row[0],
                                'rank': row[2] if len(row) > 2 else 'N/A',
                                'type': row[6] if len(row) > 6 else 'N/A',
                                'raw_data': row
                            }
                            break
                except:
                    continue
            return results
        else:
            if ranking_type not in file_map:
                return []
            
            file_path = os.path.join(DATA_DIR, f'5_参考数据/{file_map[ranking_type]}')
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb[wb.sheetnames[0]]
            
            data = []
            for row in ws.iter_rows(values_only=True):
                if row[0] is not None:
                    if school_name and school_name not in str(row[0]):
                        continue
                    data.append(row)
                    if len(data) >= 15:
                        break
            
            return data
    
    @staticmethod
    def analyze_school_comprehensive_ranking(school_name):
        """综合分析院校排名"""
        rankings = SchoolInfoSkill.get_school_ranking(school_name=school_name, ranking_type='综合')
        
        if not rankings:
            return {'school': school_name, 'error': '未找到排名数据'}
        
        avg_rank = []
        rank_details = []
        for r_type, info in rankings.items():
            rank = info['rank']
            if isinstance(rank, (int, float)) and rank > 0:
                avg_rank.append(rank)
                rank_details.append(f"{r_type}: {rank}")
        
        if avg_rank:
            average = sum(avg_rank) / len(avg_rank)
            return {
                'school': school_name,
                'rank_details': rank_details,
                'avg_rank': round(average, 1),
                'best_rank': min(avg_rank),
                'worst_rank': max(avg_rank),
                'rank_types_count': len(rankings)
            }
        
        return {'school': school_name, 'error': '无法计算综合排名'}
    
    @staticmethod
    def get_discipline_evaluation(school_name=None, discipline=None):
        """获取学科评估结果"""
        data = []
        for row in read_rows('5_参考数据/第四轮学科评估.xlsx'):
            if not clean_text(row.get('学校名称')):
                continue
            if school_name and school_name not in clean_text(row.get('学校名称')):
                continue
            if discipline and discipline not in clean_text(row.get('专业类')):
                continue
            data.append(row)
            if len(data) >= 15:
                break
        return data
    
    @staticmethod
    def analyze_school_disciplines(school_name):
        """分析院校学科实力"""
        disciplines = SchoolInfoSkill.get_discipline_evaluation(school_name=school_name)
        
        if not disciplines:
            return {'school': school_name, 'error': '未找到学科评估数据'}
        
        a_count = 0
        b_count = 0
        c_count = 0
        discipline_list = []
        
        for row in disciplines:
            level = clean_text(row.get('评估结果'))
            discipline_name = clean_text(row.get('专业类'), '未知')
            discipline_list.append(f"{discipline_name}: {level}")
            if 'A' in level:
                a_count += 1
            elif 'B' in level:
                b_count += 1
            elif 'C' in level:
                c_count += 1
        
        return {
            'school': school_name,
            'total_disciplines': len(discipline_list),
            'A_level': a_count,
            'B_level': b_count,
            'C_level': c_count,
            'discipline_list': discipline_list[:10],
            'discipline_strength': '强' if a_count >= 3 else '较强' if a_count >= 1 else '一般'
        }
    
    @staticmethod
    def compare_schools(schools):
        """对比多所院校"""
        comparison = {}
        
        for school in schools:
            ranking = SchoolInfoSkill.analyze_school_comprehensive_ranking(school)
            disciplines = SchoolInfoSkill.analyze_school_disciplines(school)
            
            comparison[school] = {
                'ranking': ranking,
                'disciplines': disciplines
            }
        
        return comparison
    
    @staticmethod
    def get_b保研_info(school_name=None):
        """获取保研资格院校信息"""
        file_path = os.path.join(DATA_DIR, '5_参考数据/保研资格院校保研率统计-2021.xls')
        wb = xlrd.open_workbook(file_path)
        ws = wb.sheet_by_index(0)
        
        data = []
        for i in range(min(ws.nrows, 50)):
            row = ws.row_values(i)
            if row[0] != '':
                if school_name and school_name not in str(row[0]):
                    continue
                data.append(tuple(row))
                if len(data) >= 10:
                    break
        
        return data
    
    @staticmethod
    def analyze_school_b保研_rate(school_name):
        """分析院校保研率"""
        raw_data = SchoolInfoSkill.get_b保研_info(school_name=school_name)
        
        if not raw_data:
            return {'school': school_name, 'error': '未找到保研数据'}
        
        for row in raw_data:
            if school_name in str(row[0]):
                return {
                    'school': row[0],
                    '保研率': row[1] if len(row) > 1 else 'N/A',
                    '保研资格': '有' if len(row) > 2 and str(row[2]) == '是' else '无',
                    'raw_data': row
                }
        
        return {'school': school_name, 'error': '未找到具体保研率数据'}

SKILLS_MAP = {
    'get_school_info': SchoolInfoSkill.get_school_info,
    'get_school_detailed_info': SchoolInfoSkill.get_school_detailed_info,
    'get_school_ranking': SchoolInfoSkill.get_school_ranking,
    'analyze_school_comprehensive_ranking': SchoolInfoSkill.analyze_school_comprehensive_ranking,
    'get_discipline_evaluation': SchoolInfoSkill.get_discipline_evaluation,
    'analyze_school_disciplines': SchoolInfoSkill.analyze_school_disciplines,
    'compare_schools': SchoolInfoSkill.compare_schools,
    'get_b保研_info': SchoolInfoSkill.get_b保研_info,
    'analyze_school_b保研_rate': SchoolInfoSkill.analyze_school_b保研_rate,
    'get_school_campus_info': SchoolInfoSkill.get_school_campus_info,
    'analyze_school_medical_strength': SchoolInfoSkill.analyze_school_medical_strength,
}

SKILL_DESCRIPTIONS = {
    'get_school_info': '获取院校基础信息（地址、电话等）',
    'get_school_detailed_info': '获取院校详细信息（含保研率、硕士点、博士点等）',
    'get_school_ranking': '获取院校排名（支持多种排名类型）',
    'analyze_school_comprehensive_ranking': '综合分析院校排名（多排名体系加权）',
    'get_discipline_evaluation': '获取学科评估结果',
    'analyze_school_disciplines': '分析院校学科实力（A/B/C等级分布）',
    'compare_schools': '对比多所院校的综合实力',
    'get_b保研_info': '获取保研资格院校信息',
    'analyze_school_b保研_rate': '分析院校保研率',
    'get_school_campus_info': '获取院校校区信息',
    'analyze_school_medical_strength': '分析院校医学学科实力',
}
